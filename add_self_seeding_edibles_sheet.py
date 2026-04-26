from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"G:\My Drive\Home Docs\Garden\Zone_7_Perennial_Subsistence_Crops__comprehensive_list_.xlsx")
SHEET_NAME = "Zone_7_Self_Seeding_Edibles"

CATEGORY_ORDER = [
    "Roots, tubers & rhizomes",
    "Leafy greens",
    "Shoots, stalks & flower vegetables",
    "Alliums",
    "Culinary herbs",
    "Tea, spice & sweetener herbs",
    "Legumes & protein crops",
    "Vines & climbers",
    "Aquatic & wetland edibles",
    "Grains, seeds & pseudocereals",
    "Succulent/xeric edibles",
]


def note_phrase_for_score(score, category):
    if score == 1:
        return "mainly a survival food or only marginally palatable"
    if score == 2:
        return "mainly considered useful rather than notably tasty"
    if score == 3:
        return "opinions are mixed or it is only good in certain preparations"
    if score == 4:
        if category in {"Culinary herbs", "Tea, spice & sweetener herbs", "Alliums"}:
            return "generally regarded as tasty in culinary use"
        return "generally regarded as tasty"
    if score == 5:
        if category in {"Culinary herbs", "Tea, spice & sweetener herbs", "Alliums", "Leafy greens", "Shoots, stalks & flower vegetables"}:
            return "widely regarded as highly tasty in its culinary role"
        return "widely regarded as highly tasty"
    return ""


def build_notes(category, tastiness_score, status, flavor, best_prep="", cautions="", extra=""):
    parts = [f"Status: {status}.", f"Tastiness: {note_phrase_for_score(tastiness_score, category)}."]
    if flavor:
        parts.append(f"Flavor: {flavor}.")
    if best_prep:
        parts.append(f"Best prep: {best_prep}.")
    if cautions:
        parts.append(f"Cautions: {cautions}.")
    if extra:
        parts.append(f"Notes: {extra}.")
    return " ".join(parts)


def record(
    category,
    common_name,
    scientific_name,
    growth_form,
    hardiness_range,
    zone_7_fit,
    sun,
    water,
    harvest_window,
    harvest_difficulty,
    tastiness,
    productivity,
    edible_parts,
    status,
    flavor,
    best_prep,
    propagation,
    nitrogen_fixer,
    cautions="",
    extra="",
):
    return {
        "Category": category,
        "Common name": common_name,
        "Scientific name": scientific_name,
        "Growth form": growth_form,
        "Will you grow it?": "",
        "Hardiness range": hardiness_range,
        "Zone 7 fit": zone_7_fit,
        "Sun": sun,
        "Water": water,
        "Harvest window": harvest_window,
        "Harvest difficulty (1-5)": harvest_difficulty,
        "Tastiness (1-5)": tastiness,
        "Productivity (1-5)": productivity,
        "Edible parts / uses": edible_parts,
        "Notes": build_notes(category, tastiness, status, flavor, best_prep, cautions, extra),
        "Propagation": propagation,
        "Nitrogen fixer": nitrogen_fixer,
    }


RECORDS = [
    record("Roots, tubers & rhizomes", "Burdock / greater burdock", "Arctium lappa", "Biennial", "3-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Fall - spring (year 1 roots)", 3, 3, 2, "Roots; peeled leaf stalks", "wild naturalized and cultivated", "earthy, sweet, artichoke-like root", "dig first-year roots; peel and simmer, roast, or stir-fry; blanch peeled stalks", "Seed", "No", "second-year roots get woody; burrs spread easily", ""),
    record("Roots, tubers & rhizomes", "Parsnip", "Pastinaca sativa", "Biennial", "3-9", "Locally reliable self-seeder", "Full sun", "Moderate", "Fall - spring (year 1 roots)", 3, 4, 3, "Roots", "cultivated and wild naturalized", "sweet, nutty, carrot-like", "dig after frost for sweeter roots; roast, mash, or add to soups", "Seed", "No", "foliage and sap can cause phytophotodermatitis", ""),
    record("Roots, tubers & rhizomes", "Wild carrot / Queen Anne's lace", "Daucus carota", "Biennial", "3-9", "Wild in zone 7", "Full sun", "Low - moderate", "Fall - spring (young roots)", 3, 2, 2, "Young roots; flowers; seeds", "wild naturalized", "carroty when very young, then fibrous", "use only first-year roots while small; seeds can be used as a spice", "Seed", "No", "dangerous lookalikes include poison hemlock", ""),
    record("Roots, tubers & rhizomes", "Salsify", "Tragopogon porrifolius", "Biennial", "4-8", "Reliable self-seeder", "Full sun", "Moderate", "Fall - spring", 3, 4, 2, "Roots; shoots; flower buds", "cultivated and wild naturalized", "mild, sweet, oyster-like root", "dig in the first year; peel roots and cook gently or use young shoots", "Seed", "No", "", ""),
    record("Roots, tubers & rhizomes", "Evening primrose", "Oenothera biennis", "Biennial", "4-9", "Wild in zone 7", "Full sun", "Low - moderate", "Fall roots; spring shoots", 3, 3, 2, "Roots; young leaves; flower buds", "wild native and cultivated", "peppery greens and mildly sweet root", "use first-year roots cooked; harvest young greens before they toughen", "Seed", "No", "", "Often persists by reseeding in open ground."),

    record("Leafy greens", "Garlic mustard", "Alliaria petiolata", "Biennial", "5-8", "Volunteer risk", "Part shade - shade", "Moderate", "Spring leaves; spring flowers; early summer pods", 1, 3, 4, "Leaves; flowers; seed pods; roots", "wild naturalized", "garlicky, mustardy, sometimes bitter with age", "use very young leaves for pesto or saute; flowers and tender pods are milder than mature leaves", "Seed", "No", "highly invasive in eastern woodlands", ""),
    record("Leafy greens", "Lamb's quarters", "Chenopodium album", "Annual", "3-10", "Volunteer risk", "Full sun", "Low - moderate", "Spring - fall", 1, 4, 5, "Leaves; young tips; seeds", "wild naturalized and cultivated", "mild, spinach-like", "use young leaves fresh or cooked; seeds can be cleaned and cooked like a pseudograin", "Seed", "No", "contains oxalates; excess nitrates can accumulate in rich soils", ""),
    record("Leafy greens", "Magenta spreen / tree spinach", "Chenopodium giganteum", "Annual", "4-9", "Locally reliable self-seeder", "Full sun - part shade", "Moderate", "Spring - fall", 1, 4, 5, "Leaves; shoots; seeds", "cultivated", "mild, spinach-like", "harvest repeatedly as cut-and-come-again greens; older leaves cook best", "Seed", "No", "contains oxalates", ""),
    record("Leafy greens", "Redroot pigweed", "Amaranthus retroflexus", "Annual", "3-10", "Wild in zone 7", "Full sun", "Low - moderate", "Summer - fall", 1, 3, 4, "Leaves; young tips; seeds", "wild naturalized", "green, earthy, similar to stronger spinach", "use young leaves cooked or wilted; seeds can be toasted or porridge-cooked", "Seed", "No", "nitrates can accumulate in stressed or heavily manured plants", ""),
    record("Leafy greens", "Chickweed", "Stellaria media", "Winter annual", "4-9", "Reliable self-seeder", "Sun - part shade", "Moderate - moist", "Cool seasons", 1, 4, 4, "Leaves; stems; flowers", "wild naturalized and cultivated", "mild, juicy, corn-silk-like", "best fresh in salads, sandwiches, or lightly wilted", "Seed", "No", "", ""),
    record("Leafy greens", "Shepherd's purse", "Capsella bursa-pastoris", "Annual/biennial", "4-9", "Wild in zone 7", "Sun - part shade", "Low - moderate", "Cool seasons", 1, 3, 3, "Leaves; young shoots; seed pods", "wild naturalized", "peppery, cress-like", "use young rosettes raw or cooked; pods can be used as a spicy garnish", "Seed", "No", "", ""),
    record("Leafy greens", "Common mallow / cheeseweed", "Malva neglecta", "Annual/biennial", "4-9", "Wild in zone 7", "Sun - part shade", "Low - moderate", "Spring - fall", 1, 3, 4, "Leaves; immature fruits", "wild naturalized", "mild and mucilaginous", "best mixed into soups, sauteed greens, or used young in salads", "Seed", "No", "texture can be slippery if overused", ""),
    record("Leafy greens", "Hairy bittercress", "Cardamine hirsuta", "Annual/biennial", "4-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Cool seasons", 1, 3, 3, "Leaves; flowers", "wild naturalized", "peppery, sharp", "use sparingly raw like cress or blend into herb sauces", "Seed", "No", "", ""),
    record("Leafy greens", "Miner's lettuce", "Claytonia perfoliata", "Annual", "5-9", "Locally reliable self-seeder", "Part shade", "Moderate", "Spring", 1, 4, 4, "Leaves; stems", "cultivated and wild naturalized", "mild, succulent, slightly sweet", "best fresh as salad greens while plants are tender", "Seed", "No", "fades quickly in heat", ""),
    record("Leafy greens", "Yellow rocket / wintercress", "Barbarea vulgaris", "Biennial/short-lived", "4-8", "Wild in zone 7", "Sun - part shade", "Moderate", "Spring", 1, 3, 3, "Young leaves; flower buds", "wild naturalized", "mustardy, sometimes bitter", "use early spring leaves and buds before flowering for the best flavor", "Seed", "No", "older leaves become bitter", ""),
    record("Leafy greens", "Corn salad / mache", "Valerianella locusta", "Winter annual", "5-8", "Reliable self-seeder", "Sun - part shade", "Moderate", "Cool seasons", 1, 4, 4, "Leaves", "cultivated and wild naturalized", "mild, nutty, tender", "harvest whole rosettes in cool weather for salads", "Seed", "No", "", ""),
    record("Leafy greens", "Garden lettuce / volunteer lettuce", "Lactuca sativa", "Annual", "5-9", "Locally reliable self-seeder", "Sun - part shade", "Moderate", "Spring - early summer; fall", 1, 4, 4, "Leaves", "cultivated", "mild to pleasantly bitter depending on strain", "use volunteers young; quality varies because seedlings may not match the parent strain", "Seed", "No", "volunteers can revert or vary in quality", ""),
    record("Leafy greens", "Wild lettuce", "Lactuca serriola", "Annual/biennial", "5-9", "Wild in zone 7", "Full sun", "Low - moderate", "Spring", 1, 2, 2, "Young leaves", "wild naturalized", "very bitter", "only the youngest leaves are worthwhile; best mixed with milder greens", "Seed", "No", "milky sap and bitterness limit use", ""),
    record("Leafy greens", "Garden orache", "Atriplex hortensis", "Annual", "4-9", "Reliable self-seeder", "Full sun", "Moderate", "Summer", 1, 4, 4, "Leaves; young tips", "cultivated", "mild, spinach-like, slightly salty", "use as cut-and-come-again greens or wilt larger leaves", "Seed", "No", "", ""),
    record("Leafy greens", "Garden arugula / rocket", "Eruca vesicaria", "Annual", "4-9", "Reliable self-seeder", "Full sun", "Moderate", "Spring - fall", 1, 4, 4, "Leaves; flowers", "cultivated and wild naturalized", "peppery, nutty", "best young in salads or quickly wilted", "Seed", "No", "older leaves get much hotter", ""),

    record("Shoots, stalks & flower vegetables", "Mustard greens", "Brassica juncea", "Annual", "5-9", "Reliable self-seeder", "Full sun", "Moderate", "Spring - fall", 1, 4, 5, "Leaves; flower shoots; young seed pods", "cultivated and wild naturalized", "mustardy and pungent", "cut repeatedly while young; flower shoots are good lightly cooked", "Seed", "No", "", ""),
    record("Shoots, stalks & flower vegetables", "Black mustard", "Brassica nigra", "Annual", "5-9", "Volunteer risk", "Full sun", "Low - moderate", "Spring - early summer", 1, 3, 4, "Leaves; flower shoots; seeds", "wild naturalized and cultivated", "strong mustard heat", "use young greens cooked or raw in small amounts; seeds can be used as spice", "Seed", "No", "can self-sow aggressively", ""),
    record("Shoots, stalks & flower vegetables", "Wild mustard / charlock", "Sinapis arvensis", "Annual", "4-9", "Wild in zone 7", "Full sun", "Low - moderate", "Spring - early summer", 1, 3, 4, "Leaves; flower buds; young seed pods", "wild naturalized", "mustardy, peppery", "harvest before flowering for milder greens; buds cook like tiny broccoli", "Seed", "No", "older plants become fibrous and hot", ""),
    record("Shoots, stalks & flower vegetables", "Radish", "Raphanus sativus", "Annual", "4-9", "Locally reliable self-seeder", "Full sun", "Moderate", "Spring - summer; fall", 2, 4, 4, "Roots; leaves; seed pods", "cultivated", "peppery roots and leaves; crisp pods", "volunteer roots are best harvested small; pods are often the best use from self-sown plants", "Seed", "No", "roots turn woody fast in heat", ""),
    record("Shoots, stalks & flower vegetables", "Wild radish", "Raphanus raphanistrum", "Annual", "4-9", "Wild in zone 7", "Full sun", "Low - moderate", "Spring - early summer", 2, 2, 3, "Leaves; flowers; pods", "wild naturalized", "peppery and often coarse", "use only the youngest greens and tender pods", "Seed", "No", "roots are usually not worthwhile", ""),
    record("Shoots, stalks & flower vegetables", "Nasturtium", "Tropaeolum majus", "Annual", "6-10", "Locally reliable self-seeder", "Full sun", "Moderate", "Summer - fall", 1, 4, 4, "Leaves; flowers; green seeds", "cultivated", "peppery, watercress-like", "use fresh leaves and flowers, or pickle the green seeds as faux capers", "Seed", "No", "volunteer behavior depends on winter and soil exposure", ""),
    record("Shoots, stalks & flower vegetables", "Mizuna / Japanese mustard greens", "Brassica rapa var. nipposinica", "Annual/biennial", "5-9", "Reliable self-seeder", "Full sun - part shade", "Moderate", "Spring - fall", 1, 4, 5, "Leaves; flower shoots", "cultivated", "mild mustard bite", "cut repeatedly while young; flowering stems are also good cooked", "Seed", "No", "", ""),
    record("Shoots, stalks & flower vegetables", "Calendula", "Calendula officinalis", "Annual", "5-10", "Reliable self-seeder", "Full sun", "Moderate", "Summer - fall", 1, 3, 4, "Petals; young leaves", "cultivated", "resinous petals, mild to bitter leaves", "use petals fresh or dried as a garnish or colorant; harvest leaves very young", "Seed", "No", "main value is culinary color and repeated bloom", ""),

    record("Culinary herbs", "Cilantro / coriander", "Coriandrum sativum", "Annual", "4-10", "Reliable self-seeder", "Full sun - part shade", "Moderate", "Spring - early summer; fall", 1, 5, 5, "Leaves; green seeds; dry seeds", "cultivated", "bright, citrusy, herbal", "use leaves young, green seeds fresh, and dry seeds as coriander spice", "Seed", "No", "", "A classic self-sowing herb in cool-season windows."),
    record("Culinary herbs", "Dill", "Anethum graveolens", "Annual", "2-10", "Reliable self-seeder", "Full sun", "Moderate", "Spring - summer", 1, 5, 4, "Leaves; flowers; green seeds; dry seeds", "cultivated", "fresh, feathery, aromatic", "harvest leaves repeatedly and allow some umbels to mature for reseeding and spice use", "Seed", "No", "", ""),
    record("Culinary herbs", "Parsley", "Petroselinum crispum", "Biennial", "5-9", "Locally reliable self-seeder", "Sun - part shade", "Moderate", "Spring - fall", 1, 5, 5, "Leaves; stems; roots in some forms", "cultivated", "clean, green, celery-like", "cut outer stems repeatedly; let some second-year plants seed to maintain the patch", "Seed", "No", "second-year plants turn coarse once flowering advances", ""),
    record("Culinary herbs", "Shiso / perilla", "Perilla frutescens", "Annual", "5-10", "Volunteer risk", "Full sun - part shade", "Moderate", "Summer - fall", 1, 4, 5, "Leaves; young seed clusters; seeds", "cultivated and wild naturalized", "distinctive cinnamon-basil-mint note", "use fresh leaves, wraps, pickles, or tempura; harvest before plants get tough", "Seed", "No", "can reseed aggressively in some gardens", ""),
    record("Culinary herbs", "Borage", "Borago officinalis", "Annual", "5-10", "Reliable self-seeder", "Full sun", "Moderate", "Summer", 1, 4, 4, "Young leaves; flowers", "cultivated and wild naturalized", "cucumber-like", "use flowers fresh and leaves very young or cooked", "Seed", "No", "older leaves are hairy and coarse", ""),
    record("Culinary herbs", "Chervil", "Anthriscus cerefolium", "Annual", "4-8", "Reliable self-seeder", "Part shade", "Moderate", "Spring; fall", 1, 4, 4, "Leaves", "cultivated", "delicate anise-parsley flavor", "best cut young in cool weather and used fresh", "Seed", "No", "fades quickly in heat", ""),
    record("Culinary herbs", "Celery", "Apium graveolens", "Biennial", "6-9", "Locally reliable self-seeder", "Full sun", "Moderate - moist", "Summer - fall", 2, 4, 3, "Leaves; stems; seeds", "cultivated", "classic celery flavor", "volunteers are often best used for leaf celery, soup greens, and seed rather than thick blanching stalks", "Seed", "No", "wild-type volunteers can be stringier than selected cultivars", ""),

    record("Tea, spice & sweetener herbs", "Caraway", "Carum carvi", "Biennial", "4-8", "Locally reliable self-seeder", "Full sun", "Moderate", "Spring leaves; summer seeds", 2, 4, 3, "Leaves; roots; seeds", "cultivated", "rye-bread, warm spice flavor", "use spring leaves fresh, first-year roots cooked, and second-year seeds dried", "Seed", "No", "", ""),
    record("Tea, spice & sweetener herbs", "German chamomile", "Matricaria chamomilla", "Annual", "4-9", "Reliable self-seeder", "Full sun", "Moderate", "Summer", 2, 4, 4, "Flowers; young leaves", "cultivated and wild naturalized", "apple-like floral tea", "pick flowers frequently and dry promptly for tea", "Seed", "No", "", ""),
    record("Tea, spice & sweetener herbs", "Nigella / black cumin", "Nigella sativa", "Annual", "5-9", "Reliable self-seeder", "Full sun", "Low - moderate", "Summer", 2, 4, 3, "Seeds; flowers", "cultivated", "onion-oregano-pepper spice note", "let pods dry fully, then thresh seeds for spice use", "Seed", "No", "", ""),

    record("Legumes & protein crops", "Crimson clover", "Trifolium incarnatum", "Annual/biennial", "5-9", "Reliable self-seeder", "Full sun", "Moderate", "Spring", 1, 2, 2, "Leaves; flowers", "cultivated and wild naturalized", "mild, green, slightly sweet blossoms", "use young leaves sparingly and flowers in teas or salads", "Seed", "Yes", "mostly a cover crop with minor food use", ""),
    record("Legumes & protein crops", "Common vetch", "Vicia sativa", "Annual", "5-9", "Wild in zone 7", "Full sun", "Moderate", "Spring", 2, 2, 2, "Young shoots; flowers", "wild naturalized and cultivated", "pea-like shoots, mild flowers", "use only young tips or flowers rather than mature dry seed", "Seed", "Yes", "mature seeds are not a preferred staple human food", ""),
    record("Legumes & protein crops", "Hairy vetch", "Vicia villosa", "Annual/biennial", "4-8", "Wild in zone 7", "Full sun", "Moderate", "Spring", 2, 1, 1, "Young shoots; flowers", "wild naturalized and cultivated", "minor, green, pea-like", "if used at all, keep to small amounts of very young shoots or flowers", "Seed", "Yes", "grown mainly as a cover crop; not usually treated as a human food staple", ""),

    record("Grains, seeds & pseudocereals", "Sunflower", "Helianthus annuus", "Annual", "4-9", "Reliable self-seeder", "Full sun", "Low - moderate", "Late summer - fall", 3, 5, 4, "Seeds; buds; petals", "cultivated and wild naturalized", "nutty seeds and artichoke-like buds", "dry heads thoroughly for seeds; immature buds can be steamed like small artichokes", "Seed", "No", "seed quality varies in volunteers from hybrid stock", ""),
    record("Grains, seeds & pseudocereals", "Grain amaranth", "Amaranthus cruentus", "Annual", "5-10", "Locally reliable self-seeder", "Full sun", "Low - moderate", "Late summer - fall", 4, 3, 4, "Seeds; leaves", "cultivated", "nutty seeds, spinach-like leaves", "cut young leaves freely or let seedheads mature and thresh for grain", "Seed", "No", "grain harvest is more laborious than leaf use", ""),

    record("Succulent/xeric edibles", "Purslane", "Portulaca oleracea", "Succulent annual", "5-10", "Volunteer risk", "Full sun", "Low", "Summer - fall", 1, 5, 5, "Leaves; stems; seed", "wild naturalized and cultivated", "lemony, juicy, slightly salty", "use fresh in salads, quick pickles, or brief sautees while stems are tender", "Seed", "No", "can spread densely in hot beds and paths", ""),
]


def copy_header_and_widths(source_ws, target_ws, headers):
    for idx, header in enumerate(headers, start=1):
        target = target_ws.cell(row=1, column=idx, value=header)
        source = source_ws.cell(row=1, column=idx)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format

        col_letter = source.column_letter
        target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

    target_ws.freeze_panes = "A2"
    target_ws.auto_filter.ref = f"A1:{target_ws.cell(1, len(headers)).column_letter}1"


def main():
    wb = load_workbook(WORKBOOK)
    main_ws = wb[wb.sheetnames[0]]
    headers = [main_ws.cell(1, i).value for i in range(1, main_ws.max_column + 1)]

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)
    copy_header_and_widths(main_ws, ws, headers)

    sorted_records = sorted(
        RECORDS,
        key=lambda item: (CATEGORY_ORDER.index(item["Category"]), item["Common name"].lower()),
    )

    for row_idx, item in enumerate(sorted_records, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item[header])

    ws.auto_filter.ref = f"A1:{ws.cell(ws.max_row, ws.max_column).column_letter}{ws.max_row}"
    wb.save(WORKBOOK)

    counts = Counter(item["Category"] for item in sorted_records)
    print(f"Saved {len(sorted_records)} records to sheet {SHEET_NAME}")
    for category in CATEGORY_ORDER:
        if counts.get(category):
            print(f"{category}: {counts[category]}")


if __name__ == "__main__":
    main()
