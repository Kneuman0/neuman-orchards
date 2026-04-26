import re
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"G:\My Drive\Home Docs\Garden\Zone_7_Perennial_Subsistence_Crops__comprehensive_list_.xlsx")
TEMP_WORKBOOK = Path.home() / "AppData" / "Local" / "Temp" / "zone7_inventory_taste_update.xlsx"
HARVEST_HEADER = "Harvest difficulty (1-5)"


def normalize(value):
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


def parse_notes(text):
    text = text or ""
    matches = list(re.finditer(r"(Status|Taste|Cautions|Notes):", text))
    if not matches:
        return {"Status": "", "Taste": "", "Cautions": "", "Notes": ""}

    parsed = {"Status": "", "Taste": "", "Cautions": "", "Notes": ""}
    for idx, match in enumerate(matches):
        key = match.group(1)
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        parsed[key] = text[start:end].strip().strip(". ")
    return parsed


TASTINESS_OVERRIDES = {
    normalize("Apios americana"): ("widely regarded as tasty", "best dug after frost, then roasted, baked, or simmered until tender"),
    normalize("Helianthus tuberosus"): ("generally regarded as tasty", "best after frost; roast, saute, or puree rather than eating huge raw portions"),
    normalize("Camassia quamash"): ("considered tasty mainly after the right preparation", "long, slow baking develops the sweetness"),
    normalize("Camassia scilloides"): ("considered tasty mainly after the right preparation", "cook thoroughly; traditional long baking improves flavor"),
    normalize("Eutrema japonicum"): ("widely regarded as tasty as a condiment", "grate fresh and use sparingly just before serving"),
    normalize("Lilium lancifolium"): ("generally regarded as tasty", "cook bulbs well; flowers are best used fresh or lightly cooked"),
    normalize("Taraxacum officinale"): ("opinions are mixed; many people like it more than they expect", "use the youngest leaves fresh or blanch/cook older leaves"),
    normalize("Cichorium intybus"): ("opinions are mixed; more valued by bitter-green fans", "use young leaves, or cook and mellow the bitterness"),
    normalize("Plantago major"): ("mainly considered useful rather than notably tasty", "harvest only very young leaves, or use the seeds"),
    normalize("Mertensia maritima"): ("widely regarded as tasty", "best fresh in small amounts for its shellfish-like flavor"),
    normalize("Phytolacca americana"): ("not generally regarded as tasty unless prepared traditionally", "only very young shoots are used, and they need repeated boiling with water changes"),
    normalize("Asclepias syriaca"): ("generally regarded as tasty when harvested at the right stage", "use very young shoots, buds, or tiny pods; boil or steam"),
    normalize("Armoracia rusticana"): ("widely regarded as tasty as a condiment", "grate fresh root and use sparingly"),
    normalize("Sassafras albidum"): ("generally regarded as tasty in traditional use", "young leaves are best dried and ground; root flavor is classic but less favored now"),
    normalize("Gaultheria procumbens"): ("opinions are mixed; some love it and some do not", "use leaves or berries as tea, candying, or flavoring rather than bulk food"),
    normalize("Ceanothus americanus"): ("pleasant but mild rather than exciting", "dry the leaves fully and use them as tea"),
    normalize("Medicago sativa"): ("mainly considered useful rather than notably tasty", "best as sprouts or very young greens"),
    normalize("Trifolium pratense"): ("mild and usable, but not usually prized for flavor", "flowers are best in tea or mixed into salads"),
    normalize("Trifolium repens"): ("mild and usable, but not usually prized for flavor", "use very young leaves or flowers in small amounts"),
    normalize("Passiflora incarnata"): ("widely regarded as tasty", "let fruit ripen fully and wrinkle slightly for best flavor"),
    normalize("Schisandra chinensis"): ("opinions are mixed; often valued more as a tonic than a dessert fruit", "dry, sweeten, or turn into tea/syrup"),
    normalize("Akebia quinata"): ("mixed; good when fully ripe but bland if immature", "eat only fully ripe pulp; avoid under-ripe fruit"),
    normalize("Pueraria montana var. lobata"): ("mainly considered useful rather than notably tasty", "roots are best processed into starch; flowers are better for jelly or syrup"),
    normalize("Mespilus germanica"): ("generally regarded as tasty after proper ripening", "blet until soft before eating"),
    normalize("Diospyros virginiana"): ("widely regarded as tasty when dead ripe", "wait until fully soft; underripe fruit is strongly astringent"),
    normalize("Diospyros kaki"): ("widely regarded as tasty", "eat when the cultivar is at proper ripeness for its astringency type"),
    normalize("Ficus carica"): ("widely regarded as tasty", "let fruit soften fully on the plant"),
    normalize("Cydonia oblonga"): ("fragrant and good, but usually not best raw", "cook into preserves, paste, or baked dishes"),
    normalize("Malus coronaria"): ("mixed raw, but very good for jelly, cider, and sauce", "cook, sweeten, or preserve"),
    normalize("Crataegus spp."): ("mixed; pleasant in the right species but often better processed", "use soft ripe fruit in ketchup, fruit leather, or jelly"),
    normalize("Celtis occidentalis"): ("pleasant enough but usually more of a wild snack than a standout fruit", "eat fully ripe fruit fresh or dry and grind it"),
    normalize("Punica granatum"): ("widely regarded as tasty", "best fully ripe and eaten fresh or juiced"),
    normalize("Citrus trifoliata"): ("not usually regarded as tasty fresh", "best for marmalade, flavoring, or mixed preserves"),
    normalize("Aronia melanocarpa"): ("healthy and usable, but often too astringent to be a favorite fresh", "freeze, sweeten, dry, or cook"),
    normalize("Ribes nigrum"): ("widely liked by some, too musky for others", "fresh when fully ripe or processed into jam/syrup"),
    normalize("Hippophae rhamnoides"): ("very tart but often considered tasty when sweetened", "juice, syrup, or cook with sweeteners"),
    normalize("Elaeagnus umbellata"): ("generally regarded as tasty", "best fully ripe; makes excellent sauces and fruit leather"),
    normalize("Viburnum trilobum"): ("mixed fresh, but good in sauce or jelly", "cook or sweeten"),
    normalize("Rosa spp."): ("generally regarded as tasty once prepared correctly", "remove inner hairs and use hips for tea, jam, or puree"),
    normalize("Chaenomeles speciosa"): ("not usually tasty raw, but excellent cooked", "use for jelly, paste, or stewed fruit"),
    normalize("Callicarpa americana"): ("mixed fresh; best processed", "use in jelly or syrup"),
    normalize("Juglans nigra"): ("widely regarded as tasty", "cure nuts well, then crack and use roasted or in baking"),
    normalize("Juglans cinerea"): ("widely regarded as tasty", "cure and crack after drying"),
    normalize("Quercus spp."): ("can be very good, but only after significant processing", "leach tannins first, then roast, grind, or cook"),
    normalize("Ginkgo biloba"): ("tasty to some people, but usually a specialty food rather than a universal favorite", "eat only cooked seeds in moderation"),
    normalize("Typha latifolia"): ("generally regarded as useful and sometimes tasty", "shoots are easiest; rhizomes are best dried, pounded, and washed for starch"),
    normalize("Nuphar advena"): ("mainly considered useful rather than notably tasty", "dry, roast, or otherwise process seeds/rhizomes before eating"),
    normalize("Thinopyrum intermedium"): ("pleasant enough as a grain, but valued more for utility than standout flavor", "harvest dry, thresh, winnow, then cook or mill"),
    normalize("Tripsacum dactyloides"): ("pleasant enough as a grain, but valued more for utility than standout flavor", "harvest dry seed, clean, then grind or cook"),
    normalize("Silphium integrifolium"): ("mainly considered useful rather than notably tasty", "treat like an oilseed or roast lightly"),
    normalize("Fagopyrum cymosum"): ("usable, but more niche than delicious", "cook or grind seeds; young leaves are milder"),
    normalize("Opuntia humifusa"): ("generally regarded as tasty", "de-spine pads and fruit carefully; grill, saute, or use fruit fresh"),
    normalize("Yucca filamentosa"): ("generally regarded as tasty", "flowers are best fresh or lightly cooked; immature pods cook better than mature ones"),
    normalize("Atriplex canescens"): ("usable more than craveable", "use sparingly as a salty green or grind seeds"),
    normalize("Atriplex halimus"): ("usable more than craveable", "best as a salt-seasoned green in mixed dishes"),
    normalize("Agave parryi"): ("widely regarded as tasty after roasting", "slow-roast the heart or flower stalk"),
    normalize("Houttuynia cordata"): ("opinions are strongly divided", "best used in small amounts, often chopped fine or mixed with other herbs"),
}


DIFFICULTY_OVERRIDES = {
    normalize("Helianthus tuberosus"): 2,
    normalize("Apios americana"): 3,
    normalize("Sium sisarum"): 3,
    normalize("Stachys affinis"): 2,
    normalize("Dioscorea polystachya"): 4,
    normalize("Claytonia virginica"): 4,
    normalize("Camassia quamash"): 4,
    normalize("Camassia scilloides"): 4,
    normalize("Pediomelum esculentum"): 3,
    normalize("Eutrema japonicum"): 4,
    normalize("Lilium lancifolium"): 3,
    normalize("Oxalis violacea"): 2,
    normalize("Urtica dioica"): 2,
    normalize("Laportea canadensis"): 2,
    normalize("Plantago major"): 2,
    normalize("Mertensia maritima"): 2,
    normalize("Tilia americana"): 2,
    normalize("Matteuccia struthiopteris"): 2,
    normalize("Phyllostachys aureosulcata"): 3,
    normalize("Asclepias syriaca"): 4,
    normalize("Phytolacca americana"): 5,
    normalize("Allium tricoccum"): 3,
    normalize("Armoracia rusticana"): 3,
    normalize("Acer saccharum"): 4,
    normalize("Acer negundo"): 4,
    normalize("Betula lenta"): 4,
    normalize("Gleditsia triacanthos"): 3,
    normalize("Caragana arborescens"): 3,
    normalize("Desmanthus illinoensis"): 4,
    normalize("Medicago sativa"): 2,
    normalize("Trifolium pratense"): 1,
    normalize("Trifolium repens"): 1,
    normalize("Astragalus crassicarpus"): 3,
    normalize("Smilax rotundifolia"): 3,
    normalize("Pueraria montana var. lobata"): 4,
    normalize("Malus coronaria"): 2,
    normalize("Mespilus germanica"): 2,
    normalize("Asimina triloba"): 2,
    normalize("Morus spp."): 2,
    normalize("Crataegus spp."): 3,
    normalize("Celtis occidentalis"): 2,
    normalize("Citrus trifoliata"): 3,
    normalize("Vaccinium macrocarpon"): 3,
    normalize("Sambucus canadensis"): 2,
    normalize("Hippophae rhamnoides"): 4,
    normalize("Elaeagnus umbellata"): 3,
    normalize("Rosa spp."): 3,
    normalize("Chaenomeles speciosa"): 3,
    normalize("Juglans nigra"): 4,
    normalize("Juglans cinerea"): 4,
    normalize("Carya spp."): 4,
    normalize("Pinus koraiensis"): 4,
    normalize("Pinus edulis"): 4,
    normalize("Quercus spp."): 5,
    normalize("Ginkgo biloba"): 3,
    normalize("Sagittaria latifolia"): 4,
    normalize("Typha latifolia"): 5,
    normalize("Nelumbo lutea"): 4,
    normalize("Nelumbo nucifera"): 4,
    normalize("Pontederia cordata"): 3,
    normalize("Nuphar advena"): 5,
    normalize("Thinopyrum intermedium"): 5,
    normalize("Tripsacum dactyloides"): 5,
    normalize("Helianthus maximiliani"): 3,
    normalize("Silphium integrifolium"): 4,
    normalize("Fagopyrum cymosum"): 4,
    normalize("Opuntia humifusa"): 3,
    normalize("Opuntia engelmannii"): 3,
    normalize("Yucca filamentosa"): 2,
    normalize("Agave parryi"): 4,
}


def base_tastiness(category, flavor):
    flavor_l = (flavor or "").lower()
    tasty_words = ["sweet", "rich", "nutty", "juicy", "aromatic", "fragrant", "honey", "banana", "mango", "kiwi", "berry", "cherry", "apple", "pear", "grapey", "winey", "oyster", "mint", "savory", "celery"]
    mixed_words = ["bitter", "earthy", "grassy", "tannic", "starchy", "salty", "camphor", "musky", "mild"]

    if any(word in flavor_l for word in tasty_words):
        return "generally regarded as tasty"
    if category in {"Tree fruits", "Shrub, cane & bramble fruits", "Nuts & oil crops"}:
        return "generally regarded as tasty"
    if category in {"Culinary herbs", "Alliums"}:
        return "generally regarded as tasty in culinary use"
    if category == "Tea, spice & sweetener herbs":
        return "generally regarded as pleasant in seasoning, tea, or syrup use"
    if category == "Legumes & protein crops":
        return "mainly considered useful rather than notably tasty"
    if category == "Grains, seeds & pseudocereals":
        return "mainly considered useful rather than notably tasty"
    if any(word in flavor_l for word in mixed_words):
        return "opinions are mixed"
    return "generally regarded as tasty"


def base_prep(category, edible_parts, cautions):
    edible_l = (edible_parts or "").lower()
    cautions_l = (cautions or "").lower()
    if "fruit" in edible_l or "berries" in edible_l or "hips" in edible_l:
        return "best fully ripe; fresh, dried, or preserved"
    if "nuts" in edible_l or "seeds for oil" in edible_l:
        return "dry or cure first, then crack, roast, grind, or press as needed"
    if "seeds" in edible_l and category == "Grains, seeds & pseudocereals":
        return "harvest dry, then clean, thresh, or grind before use"
    if "tubers" in edible_l or "roots" in edible_l or "bulbs" in edible_l or "rhizomes" in edible_l or "corms" in edible_l:
        return "usually best cooked, roasted, or simmered until tender"
    if "shoots" in edible_l or "leaves" in edible_l or "buds" in edible_l or "flowers" in edible_l:
        if "toxic" in cautions_l or "boil" in cautions_l:
            return "harvest at the correct stage and cook as directed"
        return "best young; use fresh, lightly cooked, or blanched depending on texture"
    if "sap" in edible_l:
        return "boil down promptly into syrup or use fresh"
    return ""


def base_difficulty(category, growth_form, edible_parts):
    edible_l = (edible_parts or "").lower()
    if category in {"Leafy greens", "Culinary herbs", "Alliums"}:
        return 1
    if category == "Shoots, stalks & flower vegetables":
        return 1
    if category == "Tree fruits":
        return 1
    if category == "Shrub, cane & bramble fruits":
        return 1
    if category == "Roots, tubers & rhizomes":
        return 3
    if category == "Tea, spice & sweetener herbs":
        return 2
    if category == "Legumes & protein crops":
        return 3
    if category == "Vines & climbers":
        return 2
    if category == "Nuts & oil crops":
        return 4
    if category == "Aquatic & wetland edibles":
        return 4
    if category == "Grains, seeds & pseudocereals":
        return 5
    if category == "Succulent/xeric edibles":
        return 2
    if growth_form == "Tree" and "fruit" in edible_l:
        return 2
    return 2


def build_notes(status, tastiness, flavor, prep, cautions, extra):
    parts = []
    if status:
        parts.append(f"Status: {status}.")
    if tastiness:
        parts.append(f"Tastiness: {tastiness}.")
    if flavor:
        parts.append(f"Flavor: {flavor}.")
    if prep:
        parts.append(f"Best prep: {prep}.")
    if cautions:
        parts.append(f"Cautions: {cautions}.")
    if extra:
        parts.append(f"Notes: {extra}.")
    return " ".join(parts)


def main():
    wb = load_workbook(WORKBOOK)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

    if HARVEST_HEADER not in headers:
        insert_at = headers.index("Harvest window") + 2
        ws.insert_cols(insert_at, 1)
        ws.cell(1, insert_at).value = HARVEST_HEADER
        headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

    idx = {header: pos + 1 for pos, header in enumerate(headers)}

    for row in range(2, ws.max_row + 1):
        common_name = ws.cell(row, idx["Common name"]).value or ""
        scientific_name = ws.cell(row, idx["Scientific name"]).value or ""
        category = ws.cell(row, idx["Category"]).value or ""
        growth_form = ws.cell(row, idx["Growth form"]).value or ""
        edible_parts = ws.cell(row, idx["Edible parts / uses"]).value or ""
        parsed = parse_notes(ws.cell(row, idx["Notes"]).value or "")

        key = normalize(scientific_name) or normalize(common_name)
        tastiness, prep = TASTINESS_OVERRIDES.get(key, (None, None))
        if not tastiness:
            tastiness = base_tastiness(category, parsed["Taste"])
        if not prep:
            prep = base_prep(category, edible_parts, parsed["Cautions"])

        difficulty = DIFFICULTY_OVERRIDES.get(key, base_difficulty(category, growth_form, edible_parts))

        ws.cell(row, idx[HARVEST_HEADER]).value = difficulty
        ws.cell(row, idx["Notes"]).value = build_notes(
            parsed["Status"],
            tastiness,
            parsed["Taste"],
            prep,
            parsed["Cautions"],
            parsed["Notes"],
        )

    wb.save(TEMP_WORKBOOK)
    print(f"Saved staged update to {TEMP_WORKBOOK}")
    print(f"Rows processed: {ws.max_row - 1}")


if __name__ == "__main__":
    main()
