from __future__ import annotations

import csv
import shutil
from collections import OrderedDict
from pathlib import Path

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
KB_ROOT = ROOT / "knowledgebase"
WORKBOOK_PATH = ROOT / "Zone_7_Perennial_Subsistence_Crops__comprehensive_list_.xlsx"
PERENNIAL_SHEET_EXPORT = "zone_7_perennial_subsistence_cr.csv"
SELF_SEEDING_SHEET_EXPORT = "zone_7_self_seeding_edibles.csv"
ANNUALS_EXPORT = "zone7_self_seeding_annuals_and_short_lived_edibles.csv"

PDF_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
    ),
    "Accept": "application/pdf,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

TOPICS = OrderedDict(
    [
        ("core_microfarm_permaculture", "01_core_microfarm_permaculture"),
        ("intensive_vegetable_systems", "02_intensive_vegetable_systems"),
        ("biochar", "03_biochar"),
        ("bokashi_em", "04_bokashi_em"),
        ("compost_soil_fertility", "05_compost_soil_fertility"),
        ("compost_tea_soil_biology", "06_compost_tea_soil_biology"),
        ("fungi_mycorrhiza_biostimulants", "07_fungi_mycorrhiza_biostimulants"),
        ("forest_gardens_agroforestry", "08_forest_gardens_agroforestry"),
        ("mid_atlantic_priority", "09_mid_atlantic_priority"),
        ("companion_planting_intercropping", "10_companion_planting_intercropping"),
        ("wild_plants_foraged_integration", "11_wild_plants_foraged_integration"),
        ("beneficial_insects_pest_indicators", "12_beneficial_insects_pest_indicators"),
        ("inventories_and_local_lists", "13_inventories_and_local_lists"),
    ]
)

RESOURCES = [
    {
        "title": "Permacultural Organic Market Gardening and Economic Performance",
        "authors": "Francois Leger, Sacha Guegan, et al.",
        "year": "2015",
        "topic": "core_microfarm_permaculture",
        "kind": "copy",
        "source_path": ROOT / "Bec-Hellouin-Farm-Profitability-study.pdf",
        "filename": "Bec-Hellouin-Farm-Profitability-study.pdf",
        "doi": "",
        "source_url": "",
        "mid_atlantic": "no",
        "why": (
            "Local anchor study. Real-world economic case study of the Bec Hellouin "
            "intensive permaculture market-garden nucleus."
        ),
    },
    {
        "title": "These kevin morel microfermes 2016",
        "authors": "Kevin Morel",
        "year": "2016",
        "topic": "core_microfarm_permaculture",
        "kind": "pdf",
        "download_url": "https://www.produire-bio.fr/wp-content/uploads/2017/10/These_kevin_morel_microfermes_2016.pdf",
        "filename": "Morel-2016-microfarm-thesis.pdf",
        "doi": "",
        "source_url": "https://www.produire-bio.fr/wp-content/uploads/2017/10/These_kevin_morel_microfermes_2016.pdf",
        "mid_atlantic": "no",
        "why": (
            "Long-form dissertation covering French organic microfarms, their design, "
            "work organization, and viability logic beyond the single Bec Hellouin report."
        ),
    },
    {
        "title": "Small can be beautiful for organic market gardens: an exploration of the economic viability of French microfarms using MERLIN",
        "authors": "Kevin Morel et al.",
        "year": "2017",
        "topic": "core_microfarm_permaculture",
        "kind": "pdf",
        "download_url": "https://cfppa-ac.scenari-community.org/FAD/moduleUC3/res/Small_can_be_beautifull_microfarmers_MOREL_et_al_2017_POST_PRINT.pdf",
        "filename": "Morel-2017-small-can-be-beautiful.pdf",
        "doi": "https://doi.org/10.1016/j.agsy.2017.08.008",
        "source_url": "https://cfppa-ac.scenari-community.org/FAD/moduleUC3/res/Small_can_be_beautifull_microfarmers_MOREL_et_al_2017_POST_PRINT.pdf",
        "mid_atlantic": "no",
        "why": (
            "Direct follow-on economic modeling of microfarm viability, useful for turning "
            "Bec Hellouin-style ideas into decision rules."
        ),
    },
    {
        "title": "Permaculture-Scientific Evidence of Principles for the Agroecological Design of Farming Systems",
        "authors": "Julius Krebs and Simone Bach",
        "year": "2018",
        "topic": "core_microfarm_permaculture",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-10-03218/article_deploy/sustainability-10-03218.pdf",
        "filename": "Krebs-Bach-2018-permaculture-scientific-evidence.pdf",
        "doi": "https://doi.org/10.3390/su10093218",
        "source_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-10-03218/article_deploy/sustainability-10-03218.pdf",
        "mid_atlantic": "no",
        "why": (
            "Bridges permaculture claims to peer-reviewed agroecological evidence, which is "
            "important for a future Codex skill."
        ),
    },
    {
        "title": "A conceptual framework for alternative farmers' strategic choices: the case of French organic market gardening microfarms",
        "authors": "Kevin Morel and Francois Leger",
        "year": "2016",
        "topic": "core_microfarm_permaculture",
        "kind": "url",
        "filename": "Morel-Leger-2016-conceptual-framework.url",
        "doi": "https://doi.org/10.1080/21683565.2016.1140695",
        "source_url": "https://orgprints.org/id/eprint/40495/",
        "mid_atlantic": "no",
        "why": (
            "Key conceptual paper on how alternative microfarmers balance ecological, social, "
            "and economic tradeoffs. Saved as link because direct PDF export was blocked."
        ),
        "abstract_note": (
            "Multiple-case study of 14 French organic microfarms. Focuses on strategic decision "
            "making, tradeoffs, and viability beyond profit maximization alone."
        ),
    },
    {
        "title": "Co-Design and Experimentation of a Prototype of Agroecological Micro-Farm Meeting the Objectives Set by Climate-Smart Agriculture",
        "authors": "A. Felix et al.",
        "year": "2023",
        "topic": "intensive_vegetable_systems",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/agriculture/agriculture-13-00159/article_deploy/agriculture-13-00159.pdf",
        "filename": "Felix-et-al-2023-agroecological-microfarm-prototype.pdf",
        "doi": "https://doi.org/10.3390/agriculture13010159",
        "source_url": "https://res.mdpi.com/d_attachment/agriculture/agriculture-13-00159/article_deploy/agriculture-13-00159.pdf",
        "mid_atlantic": "no",
        "why": (
            "Recent experimental microfarm paper with explicit design criteria, useful for the "
            "technical side of small-scale agroecological farm planning."
        ),
    },
    {
        "title": "Alternative Management Practices Improve Soil Health Indices in Intensive Vegetable Cropping Systems: A Review",
        "authors": "Charlotte E. Norris and Katelyn A. Congreves",
        "year": "2018",
        "topic": "intensive_vegetable_systems",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/environmental-science/articles/10.3389/fenvs.2018.00050/pdf",
        "filename": "Norris-Congreves-2018-intensive-vegetable-soil-health-review.pdf",
        "doi": "https://doi.org/10.3389/fenvs.2018.00050",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/environmental-science/articles/10.3389/fenvs.2018.00050/pdf",
        "mid_atlantic": "no",
        "why": (
            "High-value review for intensive vegetable production: amendments, cover crops, reduced "
            "tillage, soil biology, soil chemistry, and yield tradeoffs."
        ),
    },
    {
        "title": "Crop management, labour organization, and marketing: three key issues for improving sustainability in organic vegetable farming",
        "authors": "Mireille Navarrete, Lucie Dupre, and Claire Lamine",
        "year": "2014",
        "topic": "intensive_vegetable_systems",
        "kind": "url",
        "filename": "Navarrete-Dupre-Lamine-2014-crop-management-labour-marketing.url",
        "doi": "https://doi.org/10.1080/14735903.2014.959341",
        "source_url": "https://orgprints.org/id/eprint/41046/",
        "mid_atlantic": "no",
        "why": (
            "Important market-garden systems paper on diversification, labor, and marketing. "
            "Saved as link because the direct PDF host blocked download."
        ),
        "abstract_note": (
            "Examines how species diversification affects crop management, labor organization, "
            "skills, marketing, and farm sustainability in organic vegetable farms."
        ),
    },
    {
        "title": "Using cover crops to mitigate and adapt to climate change. A review",
        "authors": "Christoph Poeplau et al.",
        "year": "2017",
        "topic": "intensive_vegetable_systems",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s13593-016-0410-x.pdf",
        "filename": "Poeplau-et-al-2017-cover-crops-climate-review.pdf",
        "doi": "https://doi.org/10.1007/s13593-016-0410-x",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s13593-016-0410-x.pdf",
        "mid_atlantic": "no",
        "why": (
            "Cover crops are foundational to ecological farming; this review supports crop-rotation, "
            "soil protection, and adaptive management decisions."
        ),
    },
    {
        "title": "Use of Biochar in agriculture",
        "authors": "A. Vera-Ruiz and coauthors",
        "year": "2020",
        "topic": "biochar",
        "kind": "pdf",
        "download_url": "https://revistas.unal.edu.co/index.php/actabiol/article/download/79466/pdf",
        "filename": "Vera-Ruiz-et-al-2020-use-of-biochar-in-agriculture.pdf",
        "doi": "https://doi.org/10.15446/abc.v25n2.79466",
        "source_url": "https://revistas.unal.edu.co/index.php/actabiol/article/download/79466/pdf",
        "mid_atlantic": "no",
        "why": "Broad biochar reference for soil amendment effects, mechanisms, and agricultural use.",
    },
    {
        "title": "Plant growth improvement mediated by nitrate capture in co-composted biochar",
        "authors": "M. Kammann et al.",
        "year": "2015",
        "topic": "biochar",
        "kind": "pdf",
        "download_url": "https://www.nature.com/articles/srep11080.pdf",
        "filename": "Kammann-et-al-2015-co-composted-biochar.pdf",
        "doi": "https://doi.org/10.1038/srep11080",
        "source_url": "https://www.nature.com/articles/srep11080.pdf",
        "mid_atlantic": "no",
        "why": (
            "Directly relevant to integrating biochar with composting rather than treating it as an isolated input."
        ),
    },
    {
        "title": "Linking the Belowground Microbial Composition, Diversity and Activity to Soilborne Disease Suppression and Growth Promotion of Tomato Amended with Biochar",
        "authors": "A. Jaiswal et al.",
        "year": "2017",
        "topic": "biochar",
        "kind": "pdf",
        "download_url": "https://www.nature.com/articles/srep44382.pdf",
        "filename": "Jaiswal-et-al-2017-biochar-tomato-disease-suppression.pdf",
        "doi": "https://doi.org/10.1038/srep44382",
        "source_url": "https://www.nature.com/articles/srep44382.pdf",
        "mid_atlantic": "no",
        "why": "Connects biochar to microbiome shifts, disease suppression, and vegetable crop performance.",
    },
    {
        "title": "Influence of deficit irrigation and biochar amendment on growth, physiology, and yield of cucumber in West Texas",
        "authors": "A. Singh et al.",
        "year": "2025",
        "topic": "biochar",
        "kind": "pdf",
        "download_url": "https://www.nature.com/articles/s41598-025-94113-y.pdf",
        "filename": "Singh-et-al-2025-cucumber-biochar-deficit-irrigation.pdf",
        "doi": "https://doi.org/10.1038/s41598-025-94113-y",
        "source_url": "https://www.nature.com/articles/s41598-025-94113-y.pdf",
        "mid_atlantic": "no",
        "why": (
            "Recent crop-level paper showing how biochar integrates with irrigation management in practical vegetable production."
        ),
    },
    {
        "title": "Recycling Agricultural Waste to Enhance Sustainable Greenhouse Agriculture: Analyzing the Cost-Effectiveness and Agronomic Benefits of Bokashi and Biochar Byproducts as Soil Amendments in Citrus Nursery Production",
        "authors": "P. Chavez-Rico and coauthors",
        "year": "2024",
        "topic": "bokashi_em",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-16-06070/article_deploy/sustainability-16-06070.pdf",
        "filename": "Chavez-Rico-et-al-2024-bokashi-biochar-citrus.pdf",
        "doi": "https://doi.org/10.3390/su16146070",
        "source_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-16-06070/article_deploy/sustainability-16-06070.pdf",
        "mid_atlantic": "no",
        "why": "Useful modern bokashi paper because it evaluates agronomic and cost dimensions together.",
    },
    {
        "title": "Utilization of Bokashi Composting and Animal Feed Silage for Sustainable Agricultural Waste Management and Environmental Impact Analysis",
        "authors": "P. Chavez-Rico and coauthors",
        "year": "2023",
        "topic": "bokashi_em",
        "kind": "pdf",
        "download_url": "https://journal.gnest.org/sites/default/files/Submissions/gnest_05214/gnest_05214_published.pdf",
        "filename": "Chavez-Rico-et-al-2023-bokashi-composting-waste-management.pdf",
        "doi": "https://doi.org/10.30955/gnj.005214",
        "source_url": "https://journal.gnest.org/sites/default/files/Submissions/gnest_05214/gnest_05214_published.pdf",
        "mid_atlantic": "no",
        "why": "Direct bokashi systems paper on fermentation-based waste handling and farm nutrient cycling.",
    },
    {
        "title": "Microbes as Biofertilizers, a Potential Approach for Sustainable Crop Production",
        "authors": "M. N. Fasusi et al.",
        "year": "2021",
        "topic": "bokashi_em",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-13-01868/article_deploy/sustainability-13-01868.pdf",
        "filename": "Fasusi-et-al-2021-microbes-as-biofertilizers.pdf",
        "doi": "https://doi.org/10.3390/su13041868",
        "source_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-13-01868/article_deploy/sustainability-13-01868.pdf",
        "mid_atlantic": "no",
        "why": "Broad microbial inoculant reference that supports EM-style thinking with a wider scientific base.",
    },
    {
        "title": "The potential of lactic acid bacteria in mediating the control of plant diseases and plant growth stimulation in crop production - A mini review",
        "authors": "A. Y. M. Abdelrahman et al.",
        "year": "2023",
        "topic": "bokashi_em",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2022.1047945/pdf",
        "filename": "Abdelrahman-et-al-2023-lactic-acid-bacteria-mini-review.pdf",
        "doi": "https://doi.org/10.3389/fpls.2022.1047945",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2022.1047945/pdf",
        "mid_atlantic": "no",
        "why": "Relevant to bokashi fermentation because lactic acid bacteria are a core functional group in those systems.",
    },
    {
        "title": "Waste Management through Composting: Challenges and Potentials",
        "authors": "T. M. Awasthi et al.",
        "year": "2020",
        "topic": "compost_soil_fertility",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-12-04456/article_deploy/sustainability-12-04456.pdf",
        "filename": "Awasthi-et-al-2020-waste-management-through-composting.pdf",
        "doi": "https://doi.org/10.3390/su12114456",
        "source_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-12-04456/article_deploy/sustainability-12-04456.pdf",
        "mid_atlantic": "no",
        "why": "Composting foundation paper on process control, constraints, and scaling limits.",
    },
    {
        "title": "Thermophilic bacteria and their thermozymes in composting processes: a review",
        "authors": "A. M. Abakari et al.",
        "year": "2023",
        "topic": "compost_soil_fertility",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1186/s40538-023-00381-z.pdf",
        "filename": "Abakari-et-al-2023-thermophilic-bacteria-composting-review.pdf",
        "doi": "https://doi.org/10.1186/s40538-023-00381-z",
        "source_url": "https://link.springer.com/content/pdf/10.1186/s40538-023-00381-z.pdf",
        "mid_atlantic": "no",
        "why": "Useful for process-level compost science: heat, decomposition, thermophiles, and microbial function.",
    },
    {
        "title": "Sustainable Agriculture Through Compost Tea: Production, Application, and Impact on Horticultural Crops",
        "authors": "G. Pane and coauthors",
        "year": "2025",
        "topic": "compost_tea_soil_biology",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/horticulturae/horticulturae-11-00433/article_deploy/horticulturae-11-00433.pdf",
        "filename": "Pane-et-al-2025-compost-tea-review.pdf",
        "doi": "https://doi.org/10.3390/horticulturae11040433",
        "source_url": "https://res.mdpi.com/d_attachment/horticulturae/horticulturae-11-00433/article_deploy/horticulturae-11-00433.pdf",
        "mid_atlantic": "no",
        "why": "Dedicated compost tea review focused on production methods, application logic, and horticultural outcomes.",
    },
    {
        "title": "Advances in plant growth-promoting bacterial inoculant technology: formulations and practical perspectives (1998-2013)",
        "authors": "A. Bashan et al.",
        "year": "2013",
        "topic": "compost_tea_soil_biology",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s11104-013-1956-x.pdf",
        "filename": "Bashan-et-al-2013-pgpr-inoculant-technology.pdf",
        "doi": "https://doi.org/10.1007/s11104-013-1956-x",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s11104-013-1956-x.pdf",
        "mid_atlantic": "no",
        "why": "Practical inoculant paper for turning microbial theory into deployment choices.",
    },
    {
        "title": "The Role of Soil Microorganisms in Plant Mineral Nutrition-Current Knowledge and Future Directions",
        "authors": "A. Jacoby et al.",
        "year": "2017",
        "topic": "compost_tea_soil_biology",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2017.01617/pdf",
        "filename": "Jacoby-et-al-2017-soil-microorganisms-plant-nutrition.pdf",
        "doi": "https://doi.org/10.3389/fpls.2017.01617",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2017.01617/pdf",
        "mid_atlantic": "no",
        "why": "Soil biology backbone paper on nutrient mobilization, uptake, and microbe-mediated fertility.",
    },
    {
        "title": "Role of Arbuscular Mycorrhizal Fungi in Plant Growth Regulation: Implications in Abiotic Stress Tolerance",
        "authors": "M. Begum et al.",
        "year": "2019",
        "topic": "fungi_mycorrhiza_biostimulants",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2019.01068/pdf",
        "filename": "Begum-et-al-2019-amf-growth-regulation-review.pdf",
        "doi": "https://doi.org/10.3389/fpls.2019.01068",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2019.01068/pdf",
        "mid_atlantic": "no",
        "why": "Core review on mycorrhizae, stress tolerance, and plant performance.",
    },
    {
        "title": "Agricultural uses of plant biostimulants",
        "authors": "P. du Jardin",
        "year": "2014",
        "topic": "fungi_mycorrhiza_biostimulants",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s11104-014-2131-8.pdf",
        "filename": "duJardin-2014-agricultural-uses-of-biostimulants.pdf",
        "doi": "https://doi.org/10.1007/s11104-014-2131-8",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s11104-014-2131-8.pdf",
        "mid_atlantic": "no",
        "why": "Important framework paper for inoculants, extracts, humic substances, and related biological amendments.",
    },
    {
        "title": "The rhizosphere: a playground and battlefield for soilborne pathogens and beneficial microorganisms",
        "authors": "J. M. Raaijmakers et al.",
        "year": "2008",
        "topic": "fungi_mycorrhiza_biostimulants",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s11104-008-9568-6.pdf",
        "filename": "Raaijmakers-et-al-2008-rhizosphere-playground-and-battlefield.pdf",
        "doi": "https://doi.org/10.1007/s11104-008-9568-6",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s11104-008-9568-6.pdf",
        "mid_atlantic": "no",
        "why": "Foundational rhizosphere paper for disease suppression, antagonism, and beneficial microbe ecology.",
    },
    {
        "title": "Temperate Agroforestry: How Forest Garden Systems Combined with People-Based Ethics Can Transform Culture",
        "authors": "N. M. Castro et al.",
        "year": "2018",
        "topic": "forest_gardens_agroforestry",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-10-02246/article_deploy/sustainability-10-02246.pdf",
        "filename": "Castro-et-al-2018-temperate-agroforestry-forest-gardens.pdf",
        "doi": "https://doi.org/10.3390/su10072246",
        "source_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-10-02246/article_deploy/sustainability-10-02246.pdf",
        "mid_atlantic": "no",
        "why": "Direct forest-garden paper and the clearest bridge from permaculture language to temperate agroforestry language.",
    },
    {
        "title": "A systematic scoping literature review into temperate food forests: an overview of the current knowledge and a research agenda",
        "authors": "N. M. Castro et al.",
        "year": "2025",
        "topic": "forest_gardens_agroforestry",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s10457-025-01327-0.pdf",
        "filename": "Castro-et-al-2025-temperate-food-forests-scoping-review.pdf",
        "doi": "https://doi.org/10.1007/s10457-025-01327-0",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s10457-025-01327-0.pdf",
        "mid_atlantic": "no",
        "why": "Recent state-of-the-literature review specifically on temperate food forests.",
    },
    {
        "title": "Home gardens: a promising approach to enhance household food security and wellbeing",
        "authors": "R. Galhena et al.",
        "year": "2013",
        "topic": "forest_gardens_agroforestry",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1186/2048-7010-2-8.pdf",
        "filename": "Galhena-et-al-2013-home-gardens-food-security.pdf",
        "doi": "https://doi.org/10.1186/2048-7010-2-8",
        "source_url": "https://link.springer.com/content/pdf/10.1186/2048-7010-2-8.pdf",
        "mid_atlantic": "no",
        "why": "Home-garden review is useful because many forest-garden systems borrow the same multistrata logic.",
    },
    {
        "title": "Agroforestry and Biodiversity",
        "authors": "P. Jose",
        "year": "2019",
        "topic": "forest_gardens_agroforestry",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-11-02879/article_deploy/sustainability-11-02879.pdf",
        "filename": "Jose-2019-agroforestry-and-biodiversity.pdf",
        "doi": "https://doi.org/10.3390/su11102879",
        "source_url": "https://res.mdpi.com/d_attachment/sustainability/sustainability-11-02879/article_deploy/sustainability-11-02879.pdf",
        "mid_atlantic": "no",
        "why": "Broad agroforestry systems review that helps justify forest-garden multifunctionality.",
    },
    {
        "title": "A Review of the Role of Forests and Agroforestry Systems in the FAO Globally Important Agricultural Heritage Systems Programme",
        "authors": "J. S. L. Ortiz and coauthors",
        "year": "2020",
        "topic": "forest_gardens_agroforestry",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/forests/forests-11-00860/article_deploy/forests-11-00860.pdf",
        "filename": "Ortiz-et-al-2020-forests-agroforestry-giahs-review.pdf",
        "doi": "https://doi.org/10.3390/f11080860",
        "source_url": "https://res.mdpi.com/d_attachment/forests/forests-11-00860/article_deploy/forests-11-00860.pdf",
        "mid_atlantic": "no",
        "why": "Useful systems-level review on how perennial multistrata systems function across heritage farming contexts.",
    },
    {
        "title": "A review to frame the utilization of Eastern black walnut (Juglans nigra L.) cultivars in alley cropping systems",
        "authors": "Benjamin Bishop, Nicholas A. Meier, Mark V. Coggeshall, Sarah T. Lovell",
        "year": "2023",
        "topic": "forest_gardens_agroforestry",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s10457-023-00909-0.pdf",
        "filename": "Bishop-et-al-2023-eastern-black-walnut-alley-cropping-review.pdf",
        "doi": "https://doi.org/10.1007/s10457-023-00909-0",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s10457-023-00909-0.pdf",
        "mid_atlantic": "no",
        "why": "Climate-relevant agroforestry review for eastern North American tree-crop systems.",
    },
    {
        "title": "How sustainable agriculture can address the environmental and human health harms of industrial agriculture",
        "authors": "Leo Horrigan, Robert S. Lawrence, and Polly Walker",
        "year": "2002",
        "topic": "mid_atlantic_priority",
        "kind": "url",
        "filename": "Horrigan-Lawrence-Walker-2002-sustainable-agriculture-harms.url",
        "doi": "https://doi.org/10.1289/ehp.02110445",
        "source_url": "https://pubmed.ncbi.nlm.nih.gov/12003747/",
        "mid_atlantic": "yes",
        "why": "Baltimore-based review that frames why agroecological redesign matters at all. Saved as link because PMC blocked direct PDF export here.",
        "abstract_note": (
            "Johns Hopkins review on fossil fuel use, soil depletion, biodiversity loss, pollution, "
            "livestock feed inefficiency, and health impacts of industrial agriculture."
        ),
    },
    {
        "title": "Long-term economic performance of organic and conventional field crops in the mid-Atlantic region",
        "authors": "Michel A. Cavigelli et al.",
        "year": "2009",
        "topic": "mid_atlantic_priority",
        "kind": "url",
        "filename": "Cavigelli-et-al-2009-mid-atlantic-organic-economics.url",
        "doi": "https://doi.org/10.1017/S1742170509002555",
        "source_url": "https://www.ars.usda.gov/research/publications/publication/?seqNo115=221045",
        "mid_atlantic": "yes",
        "why": "Beltsville / University of Maryland long-term economics paper. Saved as link because direct PDF export was blocked.",
        "abstract_note": (
            "USDA-ARS Beltsville Farming Systems Project paper comparing enterprise budgets and full-rotation "
            "economics of organic and conventional field systems in the mid-Atlantic."
        ),
    },
    {
        "title": "Biomass Production and Nitrogen Accumulation by Hairy Vetch-Cereal Rye Mixtures: A Meta-Analysis",
        "authors": "R. Thapa et al.",
        "year": "2018",
        "topic": "mid_atlantic_priority",
        "kind": "pdf",
        "download_url": "https://www.ars.usda.gov/ARSUserFiles/3122/ThapaEtAl2018.pdf",
        "filename": "Thapa-et-al-2018-hairy-vetch-cereal-rye-meta-analysis.pdf",
        "doi": "https://doi.org/10.2134/agronj2017.09.0544",
        "source_url": "https://www.ars.usda.gov/ARSUserFiles/3122/ThapaEtAl2018.pdf",
        "mid_atlantic": "yes",
        "why": "University of Maryland / Beltsville-linked cover-crop meta-analysis, directly useful for your climate and soil-building plans.",
    },
    {
        "title": "Overcoming Weed Management Challenges in Cover Crop-Based Organic Rotational No-Till Soybean Production in the Eastern United States",
        "authors": "S. B. Mirsky et al.",
        "year": "2013",
        "topic": "mid_atlantic_priority",
        "kind": "url",
        "filename": "Mirsky-et-al-2013-organic-rotational-no-till-soybean.url",
        "doi": "https://doi.org/10.1614/WT-D-12-00078.1",
        "source_url": "https://doi.org/10.1614/WT-D-12-00078.1",
        "mid_atlantic": "yes",
        "why": "Eastern U.S. no-till / cover-crop systems paper relevant to reduced tillage ecological production. Saved as stable DOI link.",
        "abstract_note": "Regionally relevant organic rotational no-till paper on weed control constraints in cover crop-based systems.",
    },
    {
        "title": "Mixing plant species in cropping systems: concepts, tools and models. A review",
        "authors": "E. Malezieux, Y. Crozat, C. Dupraz, M. Laurans",
        "year": "2009",
        "topic": "companion_planting_intercropping",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1051/agro:2007057.pdf",
        "filename": "Malezieux-et-al-2009-mixing-plant-species-review.pdf",
        "doi": "https://doi.org/10.1051/agro:2007057",
        "source_url": "https://link.springer.com/content/pdf/10.1051/agro:2007057.pdf",
        "mid_atlantic": "no",
        "why": "Foundational review on species mixtures, system design logic, and modeling for diversified cropping.",
    },
    {
        "title": "Application of Trap Cropping as Companion Plants for the Management of Agricultural Pests: A Review",
        "authors": "Shovon Chandra Sarkar, Endong Wang, Shengyong Wu, Zhongren Lei",
        "year": "2018",
        "topic": "companion_planting_intercropping",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/insects/insects-09-00128/article_deploy/insects-09-00128.pdf",
        "filename": "Sarkar-et-al-2018-trap-cropping-companion-plants-review.pdf",
        "doi": "https://doi.org/10.3390/insects9040128",
        "source_url": "https://res.mdpi.com/d_attachment/insects/insects-09-00128/article_deploy/insects-09-00128.pdf",
        "mid_atlantic": "no",
        "why": "Direct companion-planting pest-management review with practical trap-crop mechanisms and case summaries.",
    },
    {
        "title": "Belowground nitrogen transfer from legumes to non-legumes under managed herbaceous cropping systems. A review",
        "authors": "Malinda S. Thilakarathna, Michel S. McElroy, Tejendra Chapagain, Yousef A. Papadopoulos",
        "year": "2016",
        "topic": "companion_planting_intercropping",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s13593-016-0396-4.pdf",
        "filename": "Thilakarathna-et-al-2016-belowground-nitrogen-transfer-review.pdf",
        "doi": "https://doi.org/10.1007/s13593-016-0396-4",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s13593-016-0396-4.pdf",
        "mid_atlantic": "no",
        "why": "Useful for legume companion planting, relay cropping, and mixed-stand fertility design.",
    },
    {
        "title": "Combining Milpa and Push-Pull Technology for sustainable food production in smallholder agriculture. A review",
        "authors": "Felipe Libran-Embid, Adewole Olagoke, Emily A. Martin",
        "year": "2023",
        "topic": "companion_planting_intercropping",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s13593-023-00896-7.pdf",
        "filename": "Libran-Embid-et-al-2023-milpa-push-pull-review.pdf",
        "doi": "https://doi.org/10.1007/s13593-023-00896-7",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s13593-023-00896-7.pdf",
        "mid_atlantic": "no",
        "why": "Strong systems paper on companion cropping for fertility, pest management, and resilient production.",
    },
    {
        "title": "Companion Plants for Aphid Pest Management",
        "authors": "Refka Ben-Issa, Laurent Gomez, Helene Gautier",
        "year": "2017",
        "topic": "companion_planting_intercropping",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/insects/insects-08-00112/article_deploy/insects-08-00112.pdf",
        "filename": "Ben-Issa-et-al-2017-companion-plants-aphid-pest-management.pdf",
        "doi": "https://doi.org/10.3390/insects8040112",
        "source_url": "https://res.mdpi.com/d_attachment/insects/insects-08-00112/article_deploy/insects-08-00112.pdf",
        "mid_atlantic": "no",
        "why": "Focused review on a concrete companion-plant mechanism: aphid disruption, attraction, and biological control support.",
    },
    {
        "title": "Intercropping-A Low Input Agricultural Strategy for Food and Environmental Security",
        "authors": "Sagar Maitra, Akbar Hossain, Marian Brestic, Milan Skalicky",
        "year": "2021",
        "topic": "companion_planting_intercropping",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/agronomy/agronomy-11-00343/article_deploy/agronomy-11-00343.pdf",
        "filename": "Maitra-et-al-2021-intercropping-low-input-strategy-review.pdf",
        "doi": "https://doi.org/10.3390/agronomy11020343",
        "source_url": "https://res.mdpi.com/d_attachment/agronomy/agronomy-11-00343/article_deploy/agronomy-11-00343.pdf",
        "mid_atlantic": "no",
        "why": "Broad synthesis of intercropping benefits, constraints, metrics, and ecological mechanisms.",
    },
    {
        "title": "Traditional knowledge of wild edible plants used in the northwest of the Iberian Peninsula (Spain and Portugal): a comparative study",
        "authors": "Manuel Pardo-de-Santayana, Javier Tardio, Emilio Blanco, Ana Maria Carvalho",
        "year": "2007",
        "topic": "wild_plants_foraged_integration",
        "kind": "pdf",
        "download_url": "https://ethnobiomed.biomedcentral.com/counter/pdf/10.1186/1746-4269-3-27",
        "filename": "Pardo-de-Santayana-et-al-2007-wild-edible-plants-iberian-peninsula.pdf",
        "doi": "https://doi.org/10.1186/1746-4269-3-27",
        "source_url": "https://ethnobiomed.biomedcentral.com/counter/pdf/10.1186/1746-4269-3-27",
        "mid_atlantic": "no",
        "why": "Strong ethnobotanical baseline on wild edible species, edible parts, and knowledge systems around gathered foods.",
    },
    {
        "title": "Open-source food: Nutrition, toxicology, and availability of wild edible greens in the East Bay",
        "authors": "Philip B. Stark, Daphne Miller, Thomas J. Carlson, Kristen Rasmussen de Vasquez",
        "year": "2019",
        "topic": "wild_plants_foraged_integration",
        "kind": "pdf",
        "download_url": "https://journals.plos.org/plosone/article/file?id=10.1371/journal.pone.0202450&type=printable",
        "filename": "Stark-et-al-2019-open-source-food-wild-edible-greens.pdf",
        "doi": "https://doi.org/10.1371/journal.pone.0202450",
        "source_url": "https://journals.plos.org/plosone/article/file?id=10.1371/journal.pone.0202450&type=printable",
        "mid_atlantic": "no",
        "why": "Useful for edible weeds as a managed resource: nutrient density, toxicology, and abundance in disturbed urban ecologies.",
    },
    {
        "title": "Useful Plants from the Wild to Home Gardens: An Analysis of Home Garden Ethnobotany in Contexts of Habitat Conversion and Land Use Change in Jeju, South Korea",
        "authors": "Yooinn Hong and Karl S. Zimmerer",
        "year": "2022",
        "topic": "wild_plants_foraged_integration",
        "kind": "url",
        "filename": "Hong-Zimmerer-2022-useful-plants-wild-to-home-gardens.url",
        "doi": "https://doi.org/10.2993/0278-0771-42.3.6",
        "source_url": "https://doi.org/10.2993/0278-0771-42.3.6",
        "mid_atlantic": "no",
        "why": "Most direct paper in the set on moving useful wild plants into managed home-garden systems.",
        "abstract_note": (
            "Analyzes how habitat conversion and encounterability of useful wild plants can drive transplantation, tolerance, and home-garden cultivation."
        ),
    },
    {
        "title": "Edible green infrastructure: An approach and review of provisioning ecosystem services and disservices in urban environments",
        "authors": "Alessio Russo, Francisco J. Escobedo, Giuseppe T. Cirella, Stefan Zerbe",
        "year": "2017",
        "topic": "wild_plants_foraged_integration",
        "kind": "url",
        "filename": "Russo-et-al-2017-edible-green-infrastructure-review.url",
        "doi": "https://doi.org/10.1016/j.agee.2017.03.026",
        "source_url": "https://doi.org/10.1016/j.agee.2017.03.026",
        "mid_atlantic": "no",
        "why": "Review on incorporating edible plants into designed landscapes while accounting for provisioning services, contamination, and tradeoffs.",
        "abstract_note": (
            "Systematic review of edible green infrastructure typologies, ecosystem services, disservices, and design implications for urban and peri-urban food landscapes."
        ),
    },
    {
        "title": "The new trends in wild edible plants valorization: commercial cultivation protocols, agronomic practices and future challenges",
        "authors": "Nikolaos Polyzos, Vasiliki Liava, Vasileios Antoniadis, Pedro Garcia",
        "year": "2025",
        "topic": "wild_plants_foraged_integration",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/horticulture/articles/10.3389/fhort.2025.1638703/pdf",
        "filename": "Polyzos-et-al-2025-wild-edible-plants-valorization-review.pdf",
        "doi": "https://doi.org/10.3389/fhort.2025.1638703",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/horticulture/articles/10.3389/fhort.2025.1638703/pdf",
        "mid_atlantic": "no",
        "why": "Direct cultivation-focused review on domesticating and managing wild edible species in farming systems.",
    },
    {
        "title": "Wild edible plants for food security, dietary diversity, and nutraceuticals: a global overview of emerging research",
        "authors": "B. Mohan Kumar, Gurulingaiah Bhavya, Savitha De Britto, Sudisha Jogaiah",
        "year": "2025",
        "topic": "wild_plants_foraged_integration",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/sustainable-food-systems/articles/10.3389/fsufs.2025.1686446/pdf",
        "filename": "Kumar-et-al-2025-wild-edible-plants-global-overview.pdf",
        "doi": "https://doi.org/10.3389/fsufs.2025.1686446",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/sustainable-food-systems/articles/10.3389/fsufs.2025.1686446/pdf",
        "mid_atlantic": "no",
        "why": "Recent overview connecting wild edible plants to food security, local food systems, domestication, and sustainable harvesting.",
    },
    {
        "title": "Wild food plant use in 21st century Europe: the disappearance of old traditions and the search for new cuisines involving wild edibles",
        "authors": "A. Luczaj et al.",
        "year": "2012",
        "topic": "wild_plants_foraged_integration",
        "kind": "url",
        "filename": "Luczaj-et-al-2012-wild-food-plant-use-in-21st-century-europe.url",
        "doi": "https://doi.org/10.5586/asbp.2012.031",
        "source_url": "https://doi.org/10.5586/asbp.2012.031",
        "mid_atlantic": "no",
        "why": "Useful cultural and culinary context for integrating foraged plants into modern food systems, kitchens, and gardens.",
        "abstract_note": "Review on changing use patterns of wild edible plants in Europe, from traditional famine foods toward contemporary culinary rediscovery.",
    },
    {
        "title": "Habitat Management to Conserve Natural Enemies of Arthropod Pests in Agriculture",
        "authors": "Douglas A. Landis, Stephen D. Wratten, Geoff M. Gurr",
        "year": "2000",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "url",
        "filename": "Landis-Wratten-Gurr-2000-habitat-management-natural-enemies.url",
        "doi": "https://doi.org/10.1146/annurev.ento.45.1.175",
        "source_url": "https://doi.org/10.1146/annurev.ento.45.1.175",
        "mid_atlantic": "no",
        "why": "Foundational review on habitat management, insectary resources, and conservation biological control.",
        "abstract_note": "Classic review laying out how floral resources, shelter, alternative prey, and habitat structure can conserve natural enemies and reduce pest pressure.",
    },
    {
        "title": "Plant species diversity for sustainable management of crop pests and diseases in agroecosystems: a review",
        "authors": "Alain Ratnadass, Paula Fernandes, Jacques Avelino, Robert Habib",
        "year": "2011",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://link.springer.com/content/pdf/10.1007/s13593-011-0022-4.pdf",
        "filename": "Ratnadass-et-al-2011-plant-species-diversity-pests-diseases-review.pdf",
        "doi": "https://doi.org/10.1007/s13593-011-0022-4",
        "source_url": "https://link.springer.com/content/pdf/10.1007/s13593-011-0022-4.pdf",
        "mid_atlantic": "no",
        "why": "Best broad review in the set for why plant diversity can suppress pests through resource dilution, natural enemies, microclimate, and improved crop physiology.",
    },
    {
        "title": "In-Field Habitat Management to Optimize Pest Control of Novel Soil Communities in Agroecosystems",
        "authors": "Kirsten Pearsons and John Tooker",
        "year": "2017",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/insects/insects-08-00082/article_deploy/insects-08-00082.pdf",
        "filename": "Pearsons-Tooker-2017-in-field-habitat-management-soil-communities.pdf",
        "doi": "https://doi.org/10.3390/insects8030082",
        "source_url": "https://res.mdpi.com/d_attachment/insects/insects-08-00082/article_deploy/insects-08-00082.pdf",
        "mid_atlantic": "no",
        "why": "Directly relevant to cover crops, reduced tillage, soil biodiversity, and predator support for belowground and surface-active pest control.",
    },
    {
        "title": "Flower Strips and Their Ecological Multifunctionality in Agricultural Fields",
        "authors": "Jolanta Kowalska, Malgorzata Antkowiak, Pawel Sienkiewicz",
        "year": "2022",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/agriculture/agriculture-12-01470/article_deploy/agriculture-12-01470.pdf",
        "filename": "Kowalska-et-al-2022-flower-strips-ecological-multifunctionality.pdf",
        "doi": "https://doi.org/10.3390/agriculture12091470",
        "source_url": "https://res.mdpi.com/d_attachment/agriculture/agriculture-12-01470/article_deploy/agriculture-12-01470.pdf",
        "mid_atlantic": "no",
        "why": "Practical review on flower strips as beneficial-insect infrastructure and on the tradeoffs around strip design and management.",
    },
    {
        "title": "Impacts of Wildflower Interventions on Beneficial Insects in Fruit Crops: A Review",
        "authors": "Michelle T. Fountain",
        "year": "2022",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://res.mdpi.com/d_attachment/insects/insects-13-00304/article_deploy/insects-13-00304.pdf",
        "filename": "Fountain-2022-wildflower-interventions-beneficial-insects-review.pdf",
        "doi": "https://doi.org/10.3390/insects13030304",
        "source_url": "https://res.mdpi.com/d_attachment/insects/insects-13-00304/article_deploy/insects-13-00304.pdf",
        "mid_atlantic": "no",
        "why": "Useful review on which wildflower interventions actually help beneficial insects and how placement and management matter.",
    },
    {
        "title": "Induction of Systemic Resistance against Insect Herbivores in Plants by Beneficial Soil Microbes",
        "authors": "Md. Harun-Or Rashid and Young R. Chung",
        "year": "2017",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2017.01816/pdf",
        "filename": "Rashid-Chung-2017-systemic-resistance-insect-herbivores-beneficial-microbes.pdf",
        "doi": "https://doi.org/10.3389/fpls.2017.01816",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2017.01816/pdf",
        "mid_atlantic": "no",
        "why": "Important review for the idea that healthy rhizosphere biology can improve direct plant resistance to herbivores, not just nutrition.",
    },
    {
        "title": "Plant mineral nutrition and disease resistance: A significant linkage for sustainable crop protection",
        "authors": "Ruchi Tripathi, Rashmi Tewari, K. P. Singh, Chetan Keswani, Tatiana Minkina",
        "year": "2022",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2022.883970/pdf",
        "filename": "Tripathi-et-al-2022-plant-mineral-nutrition-disease-resistance.pdf",
        "doi": "https://doi.org/10.3389/fpls.2022.883970",
        "source_url": "https://public-pages-files-2025.frontiersin.org/journals/plant-science/articles/10.3389/fpls.2022.883970/pdf",
        "mid_atlantic": "no",
        "why": "Best fit for using disease and disorder pressure as a nutrient-balance diagnostic rather than treating all problems as spray problems.",
    },
    {
        "title": "Nutrition vs association: plant defenses are altered by arbuscular mycorrhizal fungi association not by nutritional provisioning alone",
        "authors": "Chase A. Stratton, Swayamjit Ray, Brosi A. Bradley, Jason P. Kaye, Jared G. Ali",
        "year": "2022",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://bmcplantbiol.biomedcentral.com/counter/pdf/10.1186/s12870-022-03795-3.pdf",
        "filename": "Stratton-et-al-2022-amf-association-plant-defenses.pdf",
        "doi": "https://doi.org/10.1186/s12870-022-03795-3",
        "source_url": "https://bmcplantbiol.biomedcentral.com/counter/pdf/10.1186/s12870-022-03795-3.pdf",
        "mid_atlantic": "no",
        "why": "Useful mechanistic paper showing that mycorrhiza-driven defense effects are not explained by nutrition alone.",
    },
    {
        "title": "Plant Nutrition: A tool for the management of hemipteran insect-pests-A review",
        "authors": "Vinay Singh and A. K. Sood",
        "year": "2017",
        "topic": "beneficial_insects_pest_indicators",
        "kind": "pdf",
        "download_url": "https://arccarticles.s3.amazonaws.com/webArticle/Final-attachment-published-R-1637.pdf",
        "filename": "Singh-Sood-2017-plant-nutrition-hemipteran-insect-pests-review.pdf",
        "doi": "https://doi.org/10.18805/ag.r-1637",
        "source_url": "https://arccarticles.s3.amazonaws.com/webArticle/Final-attachment-published-R-1637.pdf",
        "mid_atlantic": "no",
        "why": "Direct review on how nutrient balance shifts pressure from sap-feeding pests such as aphids, leafhoppers, whiteflies, and related groups.",
    },
]

SELF_SEEDING_MASTER_FIELDS = [
    "Category",
    "Common name",
    "Scientific name",
    "Lifecycle",
    "Status",
    "Zone 7 self-seeding fit",
    "Primary edible use",
    "Notes",
]

SELF_SEEDING_MASTER = [
    {"Category": "Roots, tubers & rhizomes", "Common name": "Burdock / greater burdock", "Scientific name": "Arctium lappa", "Lifecycle": "Biennial", "Status": "both", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Roots, peeled stalks", "Notes": "Common wild volunteer; easiest when roots are taken in year one."},
    {"Category": "Roots, tubers & rhizomes", "Common name": "Evening primrose", "Scientific name": "Oenothera biennis", "Lifecycle": "Biennial", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Roots, leaves, flower buds", "Notes": "Often behaves as a persistent self-sowing patch once established."},
    {"Category": "Roots, tubers & rhizomes", "Common name": "Parsnip", "Scientific name": "Pastinaca sativa", "Lifecycle": "Biennial", "Status": "both", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Roots, young greens", "Notes": "Needs plants left to flower in year two; wild populations also persist by reseeding."},
    {"Category": "Roots, tubers & rhizomes", "Common name": "Salsify", "Scientific name": "Tragopogon porrifolius", "Lifecycle": "Biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Roots, shoots", "Notes": "Often self-sows freely when a few seed heads are left standing."},
    {"Category": "Roots, tubers & rhizomes", "Common name": "Wild carrot / Queen Anne's lace", "Scientific name": "Daucus carota", "Lifecycle": "Biennial", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Roots, greens, seeds", "Notes": "Edible only with confident identification because of toxic Apiaceae lookalikes."},
    {"Category": "Roots, tubers & rhizomes", "Common name": "Beet / beetroot", "Scientific name": "Beta vulgaris", "Lifecycle": "Biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Roots, greens", "Notes": "Can reseed after overwintering and bolting in mild zone-7 winters."},
    {"Category": "Roots, tubers & rhizomes", "Common name": "Swiss chard / leaf beet", "Scientific name": "Beta vulgaris var. cicla", "Lifecycle": "Biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves, stalks", "Notes": "More likely to reseed where plants overwinter successfully."},
    {"Category": "Roots, tubers & rhizomes", "Common name": "Daikon / tillage radish", "Scientific name": "Raphanus sativus var. longipinnatus", "Lifecycle": "Annual to biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Roots, leaves, pods", "Notes": "Can volunteer heavily if spring seed set is allowed."},
    {"Category": "Leafy greens", "Common name": "Chickweed", "Scientific name": "Stellaria media", "Lifecycle": "Annual / winter annual", "Status": "both", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Tender greens", "Notes": "Classic cool-season self-sower in garden paths and bed edges."},
    {"Category": "Leafy greens", "Common name": "Common mallow / cheeseweed", "Scientific name": "Malva neglecta", "Lifecycle": "Annual to short-lived perennial", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Leaves, immature fruits", "Notes": "Persistent from seed in disturbed soils; texture is mucilaginous."},
    {"Category": "Leafy greens", "Common name": "Corn salad / mache", "Scientific name": "Valerianella locusta", "Lifecycle": "Annual", "Status": "both", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Salad greens", "Notes": "Very dependable cool-season self-sower if spring plants are left to seed."},
    {"Category": "Leafy greens", "Common name": "Garden arugula / rocket", "Scientific name": "Eruca vesicaria", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Peppery greens, flowers", "Notes": "One of the easiest managed self-sowers in a vegetable garden."},
    {"Category": "Leafy greens", "Common name": "Garden lettuce / volunteer lettuce", "Scientific name": "Lactuca sativa", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Greens", "Notes": "Volunteers best where some summer seed shatter is tolerated."},
    {"Category": "Leafy greens", "Common name": "Garden orache", "Scientific name": "Atriplex hortensis", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Leaves", "Notes": "Useful warm-weather spinach substitute that often returns from dropped seed."},
    {"Category": "Leafy greens", "Common name": "Garlic mustard", "Scientific name": "Alliaria petiolata", "Lifecycle": "Biennial", "Status": "wild", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Leaves, flowering tops, roots", "Notes": "Edible but invasive; best handled as a forage-and-remove species, not intentionally spread."},
    {"Category": "Leafy greens", "Common name": "Hairy bittercress", "Scientific name": "Cardamine hirsuta", "Lifecycle": "Annual / winter annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Peppery greens", "Notes": "Explosive seed dispersal makes it a reliable volunteer."},
    {"Category": "Leafy greens", "Common name": "Lamb's quarters", "Scientific name": "Chenopodium album", "Lifecycle": "Annual", "Status": "both", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Leaves, seeds", "Notes": "Excellent edible weed, but prolific enough that most gardens treat it as a managed volunteer."},
    {"Category": "Leafy greens", "Common name": "Magenta spreen / tree spinach", "Scientific name": "Chenopodium giganteum", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves", "Notes": "Less aggressive than common lamb's quarters but often reseeds in open soil."},
    {"Category": "Leafy greens", "Common name": "Miner's lettuce", "Scientific name": "Claytonia perfoliata", "Lifecycle": "Annual", "Status": "both", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Salad greens", "Notes": "Best in cool, moist conditions; can naturalize in shaded spring beds."},
    {"Category": "Leafy greens", "Common name": "Redroot pigweed", "Scientific name": "Amaranthus retroflexus", "Lifecycle": "Annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Greens, seeds", "Notes": "Common edible amaranth volunteer in disturbed summer ground."},
    {"Category": "Leafy greens", "Common name": "Shepherd's purse", "Scientific name": "Capsella bursa-pastoris", "Lifecycle": "Annual / winter annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Greens, seed pods", "Notes": "Reliable seed bank species in cool-season garden margins."},
    {"Category": "Leafy greens", "Common name": "Wild lettuce", "Scientific name": "Lactuca serriola", "Lifecycle": "Annual to biennial", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Young leaves", "Notes": "Best only as a young green; more of a forage species than a deliberate crop."},
    {"Category": "Leafy greens", "Common name": "Yellow rocket / wintercress", "Scientific name": "Barbarea vulgaris", "Lifecycle": "Biennial", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Young greens", "Notes": "Better cooked or blanched when mature because flavor intensifies."},
    {"Category": "Leafy greens", "Common name": "Spinach", "Scientific name": "Spinacia oleracea", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves", "Notes": "Most reliable where fall or very early spring crops can bolt and drop seed."},
    {"Category": "Leafy greens", "Common name": "Garden cress", "Scientific name": "Lepidium sativum", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Peppery greens", "Notes": "Fast cycle makes it easy to reseed itself repeatedly in cool weather."},
    {"Category": "Leafy greens", "Common name": "Tatsoi", "Scientific name": "Brassica rapa subsp. narinosa", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves", "Notes": "Can volunteer where spring flowering is allowed."},
    {"Category": "Leafy greens", "Common name": "Bok choy / pak choi", "Scientific name": "Brassica rapa subsp. chinensis", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves, stems", "Notes": "Volunteers best from spring seed crops, less so from fall plantings."},
    {"Category": "Leafy greens", "Common name": "Henbit", "Scientific name": "Lamium amplexicaule", "Lifecycle": "Annual / winter annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Shoots, flowers", "Notes": "Common edible volunteer in cool-season bare ground."},
    {"Category": "Leafy greens", "Common name": "Purple deadnettle", "Scientific name": "Lamium purpureum", "Lifecycle": "Annual / winter annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Shoots, leaves", "Notes": "Another common winter annual edible in zone-7 gardens."},
    {"Category": "Leafy greens", "Common name": "Cleavers", "Scientific name": "Galium aparine", "Lifecycle": "Annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Very young shoots", "Notes": "Better as a spring forage than a cultivated crop; quickly becomes rough."},
    {"Category": "Leafy greens", "Common name": "Common sow thistle", "Scientific name": "Sonchus oleraceus", "Lifecycle": "Annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Young greens", "Notes": "Useful young, especially cooked; prolific seed producer."},
    {"Category": "Leafy greens", "Common name": "New Zealand spinach", "Scientific name": "Tetragonia tetragonioides", "Lifecycle": "Tender annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Leaves", "Notes": "Warm-season self-seeding is possible after a long season but not reliable everywhere."},
    {"Category": "Leafy greens", "Common name": "Leaf amaranth", "Scientific name": "Amaranthus tricolor", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves", "Notes": "Often volunteers lightly where hot summers allow full seed maturity."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Black mustard", "Scientific name": "Brassica nigra", "Lifecycle": "Annual", "Status": "both", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Greens, flower buds, seeds", "Notes": "Easy volunteer; can become weedy if unmanaged."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Calendula", "Scientific name": "Calendula officinalis", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Petals, young leaves", "Notes": "Mainly a flower crop, but one of the best edible self-sowers."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Mizuna / Japanese mustard greens", "Scientific name": "Brassica rapa var. nipposinica", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Greens", "Notes": "Frequently volunteers after spring bolting."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Mustard greens", "Scientific name": "Brassica juncea", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Greens", "Notes": "Fast and prolific seed set makes this one of the easiest brassica volunteers."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Nasturtium", "Scientific name": "Tropaeolum majus", "Lifecycle": "Tender annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves, flowers, green seed pods", "Notes": "Reseeds best in warmer pockets and open beds."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Radish", "Scientific name": "Raphanus sativus", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Roots, greens, pods", "Notes": "Seed pods are also useful food if roots are not harvested."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Wild mustard / charlock", "Scientific name": "Sinapis arvensis", "Lifecycle": "Annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Young greens", "Notes": "A forage species rather than a planted crop."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Wild radish", "Scientific name": "Raphanus raphanistrum", "Lifecycle": "Annual", "Status": "wild", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Greens, pods", "Notes": "Common volunteer or field-edge forage plant."},
    {"Category": "Shoots, stalks & flower vegetables", "Common name": "Turnip", "Scientific name": "Brassica rapa subsp. rapa", "Lifecycle": "Biennial grown as an annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Roots, greens", "Notes": "More reliable if overwintered roots are left to flower."},
    {"Category": "Culinary herbs", "Common name": "Borage", "Scientific name": "Borago officinalis", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Leaves, flowers", "Notes": "One of the classic edible herb self-sowers."},
    {"Category": "Culinary herbs", "Common name": "Celery", "Scientific name": "Apium graveolens", "Lifecycle": "Biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaf stalks, leaves, seeds", "Notes": "Needs overwintered plants to bolt and scatter seed."},
    {"Category": "Culinary herbs", "Common name": "Chervil", "Scientific name": "Anthriscus cerefolium", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Leaves", "Notes": "Cool-season herb that often naturalizes lightly."},
    {"Category": "Culinary herbs", "Common name": "Cilantro / coriander", "Scientific name": "Coriandrum sativum", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Leaves, green seeds, dry seeds", "Notes": "A core self-sowing culinary herb in zone 7."},
    {"Category": "Culinary herbs", "Common name": "Dill", "Scientific name": "Anethum graveolens", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Leaves, flowers, seeds", "Notes": "Often returns more aggressively than planned."},
    {"Category": "Culinary herbs", "Common name": "Parsley", "Scientific name": "Petroselinum crispum", "Lifecycle": "Biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves, roots, seeds", "Notes": "Needs second-year flowering plants for reseeding."},
    {"Category": "Culinary herbs", "Common name": "Shiso / perilla", "Scientific name": "Perilla frutescens", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Leaves, flower spikes, seeds", "Notes": "Very capable self-sower once established; manage if spread is unwanted."},
    {"Category": "Culinary herbs", "Common name": "Epazote", "Scientific name": "Dysphania ambrosioides", "Lifecycle": "Annual", "Status": "both", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Leaves", "Notes": "Warm-season herb that can reseed aggressively in hot sites."},
    {"Category": "Culinary herbs", "Common name": "Summer savory", "Scientific name": "Satureja hortensis", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Leaves", "Notes": "Not as aggressive as dill or cilantro but will return in open soil."},
    {"Category": "Culinary herbs", "Common name": "Florence fennel / annual fennel", "Scientific name": "Foeniculum vulgare var. azoricum", "Lifecycle": "Annual to biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Bulbs, leaves, seeds", "Notes": "Can cross into wilder fennel behavior where seed is allowed to mature."},
    {"Category": "Tea, spice & sweetener herbs", "Common name": "Caraway", "Scientific name": "Carum carvi", "Lifecycle": "Biennial", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Seeds, roots, leaves", "Notes": "Usually needs second-year flowering to perpetuate itself."},
    {"Category": "Tea, spice & sweetener herbs", "Common name": "German chamomile", "Scientific name": "Matricaria chamomilla", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Flowers", "Notes": "One of the easiest tea herbs to naturalize from seed."},
    {"Category": "Tea, spice & sweetener herbs", "Common name": "Nigella / black cumin", "Scientific name": "Nigella sativa", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Seeds", "Notes": "Often persists in lightly disturbed beds after initial sowing."},
    {"Category": "Tea, spice & sweetener herbs", "Common name": "Anise", "Scientific name": "Pimpinella anisum", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Seeds, young leaves", "Notes": "More site-sensitive than dill or cilantro but worth tracking in warm gardens."},
    {"Category": "Legumes & protein crops", "Common name": "Common vetch", "Scientific name": "Vicia sativa", "Lifecycle": "Annual", "Status": "both", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Young tips, seeds with caution", "Notes": "Mainly relevant as a volunteer green-manure and forage species; seed use is limited."},
    {"Category": "Legumes & protein crops", "Common name": "Crimson clover", "Scientific name": "Trifolium incarnatum", "Lifecycle": "Annual / winter annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Flowers, sprouts, young leaves", "Notes": "More often used as a cover crop, but edible and highly dependable for reseeding."},
    {"Category": "Legumes & protein crops", "Common name": "Hairy vetch", "Scientific name": "Vicia villosa", "Lifecycle": "Annual / winter annual", "Status": "both", "Zone 7 self-seeding fit": "Wild in zone 7", "Primary edible use": "Young tips and flowers in limited culinary use", "Notes": "Primarily a cover crop; included because it persists well and has limited edible use."},
    {"Category": "Legumes & protein crops", "Common name": "Austrian winter pea / field pea", "Scientific name": "Pisum sativum subsp. arvense", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Shoots, pods, peas", "Notes": "Can volunteer from missed seed, especially in protected or mulched beds."},
    {"Category": "Grains, seeds & pseudocereals", "Common name": "Grain amaranth", "Scientific name": "Amaranthus cruentus", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Seeds, greens", "Notes": "Good candidate where seed heads are left in place through fall."},
    {"Category": "Grains, seeds & pseudocereals", "Common name": "Sunflower", "Scientific name": "Helianthus annuus", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Seeds, buds, petals", "Notes": "Returns well if birds and rodents do not remove all seed."},
    {"Category": "Grains, seeds & pseudocereals", "Common name": "Buckwheat", "Scientific name": "Fagopyrum esculentum", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Seeds, greens", "Notes": "Especially good at summer volunteer cycles when seed is allowed to shatter."},
    {"Category": "Grains, seeds & pseudocereals", "Common name": "Quinoa", "Scientific name": "Chenopodium quinoa", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Locally reliable self-seeder", "Primary edible use": "Seeds, young leaves", "Notes": "More variable than amaranth, but volunteers can appear in open, dry summer soils."},
    {"Category": "Grains, seeds & pseudocereals", "Common name": "Breadseed poppy", "Scientific name": "Papaver somniferum", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Seeds", "Notes": "Very dependable self-sower; legal rules and ornamental/crop context should be checked locally."},
    {"Category": "Fruiting vegetables", "Common name": "Ground cherry", "Scientific name": "Physalis pruinosa", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Fruit", "Notes": "One of the best self-sowing fruiting annuals for zone 7."},
    {"Category": "Fruiting vegetables", "Common name": "Tomatillo", "Scientific name": "Physalis philadelphica", "Lifecycle": "Annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Reliable self-seeder", "Primary edible use": "Fruit", "Notes": "Frequently volunteers around compost edges and former planting spots."},
    {"Category": "Vines & climbers", "Common name": "Malabar spinach", "Scientific name": "Basella alba", "Lifecycle": "Tender perennial grown as an annual", "Status": "cultivated", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Leaves, shoots", "Notes": "Warm-season self-seeding is possible but not dependable in every mid-Atlantic garden."},
    {"Category": "Succulent/xeric edibles", "Common name": "Purslane", "Scientific name": "Portulaca oleracea", "Lifecycle": "Annual", "Status": "both", "Zone 7 self-seeding fit": "Volunteer risk", "Primary edible use": "Succulent greens", "Notes": "Very persistent from seed in bare summer soil; one of the most useful edible volunteers."},
]


def ensure_dirs() -> None:
    if KB_ROOT.exists():
        shutil.rmtree(KB_ROOT)
    KB_ROOT.mkdir(exist_ok=True)
    for folder in TOPICS.values():
        (KB_ROOT / folder).mkdir(exist_ok=True)


def download_pdf(url: str, destination: Path) -> None:
    response = requests.get(url, headers=PDF_HEADERS, timeout=120)
    response.raise_for_status()
    if not response.content.startswith(b"%PDF"):
        raise ValueError(f"URL did not return a PDF: {url}")
    destination.write_bytes(response.content)


def copy_local(source: Path, destination: Path) -> None:
    if not source.exists():
        raise FileNotFoundError(source)
    shutil.copy2(source, destination)


def write_url_shortcut(destination: Path, url: str) -> None:
    destination.write_text(f"[InternetShortcut]\nURL={url}\n", encoding="utf-8")


def write_url_note(destination: Path, resource: dict[str, str]) -> None:
    lines = [
        f"# {resource['title']}",
        "",
        f"- Authors: {resource['authors']}",
        f"- Year: {resource['year']}",
        f"- DOI: {resource['doi'] or 'n/a'}",
        f"- URL: {resource['source_url']}",
        f"- Why selected: {resource['why']}",
    ]
    abstract_note = resource.get("abstract_note", "").strip()
    if abstract_note:
        lines.extend(["", "## Summary note", "", abstract_note])
    destination.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_bec_hellouin_summary() -> None:
    summary = """# Bec Hellouin study notes

## Why this matters

This is the anchor paper already present in the workspace. It is the clearest local starting point for the knowledge base because it ties ecological market-garden design to recorded labor, yields, and economic performance.

## Main takeaways

- The study covers the most intensive 1,000 m2 production nucleus of the farm, not the whole farm.
- Roughly 42% of the studied area was under greenhouse protection.
- Data were recorded in real farm conditions from late 2011 through early 2015.
- Gross sales rose from about 33,000 EUR in 2013 to about 57,000 EUR in 2014 on the studied intensive area.
- The report ties performance gains to better operator skill, improved tools, bed redesign for ergonomics, shorter-cycle crops, market diversification, and hotbeds that stretched the season.
- The authors explicitly warn that the 1,000 m2 area should not be treated as a complete microfarm by itself. The intensive nucleus depends on a larger whole-farm ecology that includes less-intensive areas, natural areas, buildings, and supporting elements.

## Techniques and management themes to carry into the wider knowledge base

- High crop density and spatial intensification
- Intercropping and relay cropping
- Protected cultivation and hotbeds
- Manual tools with low mechanization
- Bed layout and ergonomic redesign
- Short marketing chains and close market fit
- Ecological maturity of the whole farm, not just the vegetable beds
- Long-term soil building and ecosystem-service thinking

## Constraints and cautions

- Labor demand is high, especially during intensification phases.
- The model depends heavily on technical skill, observation, and tight management.
- Economic viability is possible, but not automatic; it depends on production per area, workflow discipline, and market quality.
"""
    (KB_ROOT / "Bec_Hellouin_summary.md").write_text(summary, encoding="utf-8")


def build_beneficial_pest_note() -> None:
    path = (
        KB_ROOT
        / TOPICS["beneficial_insects_pest_indicators"]
        / "beneficial_plants_bugs_and_pest_indicators.md"
    )
    note = """# Beneficial plants, beneficial bugs, and pest-indicator notes

## Why this section exists

The papers in this folder are useful only if they can be turned into a field-reading framework. The main practical lesson is not "spray less and hope." It is to read pest pressure as information about plant vigor, nutrient balance, habitat structure, rhizosphere function, and natural-enemy support.

## Core interpretation rule

Do not assume all pest pressure means the same thing. Some pests increase on stressed plants, some on lush vigorous tissue, and some mainly when natural enemies are missing. The literature supports using pests as indicators, but not with a one-size-fits-all rule.

## Beneficial plants worth building around

- Umbel flowers such as dill, cilantro, parsley, fennel, and caraway are repeatedly useful for hoverflies and parasitoid wasps because their flowers are accessible to small-mouthed adults.
- Quick nectar strips such as buckwheat and phacelia are useful when you need rapid floral resources close to crop beds.
- Alyssum-style low flowering strips are commonly used to support hoverflies and parasitoids near vegetable crops.
- Calendula, chamomile, and mixed daisy-family flowers can broaden bloom windows and support generalist beneficial insects.
- Mustards and nasturtiums can function as trap or diversion plants in some systems, but they must be managed so they do not simply become pest reservoirs.
- Perennial edges, clovers, cover crops, and reduced-disturbance zones support ground beetles, rove beetles, spiders, and other predators that work near or below the soil surface.

## Beneficial bugs and what they usually help with

- Hoverflies: aphids and other small soft-bodied pests. Adults need nectar and pollen; larvae do the predation.
- Lady beetles: aphids, scales, mites, and other soft-bodied prey.
- Lacewings: aphids, thrips, mites, small caterpillars, and eggs.
- Parasitic wasps: aphids, caterpillars, whiteflies, and many concealed pests, depending on species.
- Minute pirate bugs: thrips, mites, and small soft-bodied insects.
- Ground beetles and rove beetles: slugs, caterpillars, pupae, and soil-surface pests.
- Predatory mites: spider mites, thrips, and small arthropods in humid canopies.
- Tachinid flies and predatory wasps: caterpillars and larger chewing pests.

## Pest pressure as a garden diagnostic

- Aphids, whiteflies, and similar sap feeders often point to a combination of soluble nitrogen excess, water imbalance, lush weak growth, or poor natural-enemy support.
- Flea beetles and related early-season brassica pests often hit hardest when seedlings are slow, stalled, or nutritionally constrained. The fix is often faster early growth and better establishment, not only direct control.
- Repeated foliar disease and soft growth often suggest unbalanced fertility, especially too much nitrogen relative to structural and defensive support from other nutrients.
- Root-zone pest or disease pressure often points toward disturbed soil food webs, weak aggregate structure, excess moisture swings, or low biological buffering in the rhizosphere.
- Persistent outbreaks in simplified beds may indicate a habitat problem: too little flowering continuity, too few refuge zones, too much bare soil, or too much disturbance for predators and parasitoids to stay active.

## What the evidence base actually supports

- Diverse plantings can reduce pest pressure through multiple pathways at once: host dilution, olfactory and visual disruption, natural-enemy support, microclimate shifts, and improved crop physiology.
- Floral resources help beneficial insects, but not all flowers help all beneficials. Flower shape, bloom timing, strip placement, and pest-natural-enemy matching matter.
- Healthy soil biology can prime plant defenses. Beneficial microbes and mycorrhizae can shift plant defense chemistry, not just nutrient uptake.
- Balanced mineral nutrition is part of crop protection. Plants that are overfed with nitrogen or otherwise nutritionally imbalanced can become more susceptible to disease and some insect guilds.
- Soil-covering systems with cover crops, residue, and reduced tillage can strengthen predator communities, but only if moisture and disturbance are managed so the system does not simply favor the pest as well.

## Practical caution

Dense, dark-green, fast-growing plants are not automatically pest-proof. The literature contains both plant-stress and plant-vigor patterns. The right interpretation depends on the pest guild:

- Sap feeders often benefit from certain stress or nutrient-imbalance conditions.
- Gallers and some specialists may prefer vigorous tissue.
- Generalist predators usually need habitat continuity and low enough disturbance to persist.

## Most defensible garden response sequence

1. Identify the pest guild correctly.
2. Check whether crop growth is stalled, lush and weak, drought-stressed, or shaded too heavily.
3. Check nitrogen push versus mineral balance and biological support.
4. Check whether there is continuous nearby habitat for beneficials: flowers, refuge, mulch, cover, and low-spray zones.
5. Add direct controls only after the crop-health and habitat variables have been read honestly.

## Evidence base in this folder

- `Landis-Wratten-Gurr-2000-habitat-management-natural-enemies.url`
- `Ratnadass-et-al-2011-plant-species-diversity-pests-diseases-review.pdf`
- `Pearsons-Tooker-2017-in-field-habitat-management-soil-communities.pdf`
- `Kowalska-et-al-2022-flower-strips-ecological-multifunctionality.pdf`
- `Fountain-2022-wildflower-interventions-beneficial-insects-review.pdf`
- `Rashid-Chung-2017-systemic-resistance-insect-herbivores-beneficial-microbes.pdf`
- `Tripathi-et-al-2022-plant-mineral-nutrition-disease-resistance.pdf`
- `Stratton-et-al-2022-amf-association-plant-defenses.pdf`
- `Singh-Sood-2017-plant-nutrition-hemipteran-insect-pests-review.pdf`
"""
    path.write_text(note, encoding="utf-8")


def build_catalog() -> None:
    catalog_path = KB_ROOT / "catalog.csv"
    with catalog_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "topic_folder",
                "title",
                "authors",
                "year",
                "kind",
                "filename",
                "doi",
                "source_url",
                "mid_atlantic",
                "why",
            ],
        )
        writer.writeheader()
        for item in RESOURCES:
            writer.writerow(
                {
                    "topic_folder": TOPICS[item["topic"]],
                    "title": item["title"],
                    "authors": item["authors"],
                    "year": item["year"],
                    "kind": item["kind"],
                    "filename": item["filename"],
                    "doi": item["doi"],
                    "source_url": item["source_url"],
                    "mid_atlantic": item["mid_atlantic"],
                    "why": item["why"],
                }
            )


def export_sheet_to_csv(destination: Path, sheet_name: str) -> int:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[sheet_name]
    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])
    return max(0, ws.max_row - 1)


def build_inventory_assets() -> None:
    folder = KB_ROOT / TOPICS["inventories_and_local_lists"]
    copy_local(WORKBOOK_PATH, folder / WORKBOOK_PATH.name)

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet_exports: list[tuple[str, str, int]] = []
    for ws in wb.worksheets:
        safe_name = ws.title.lower().replace(" ", "_")
        destination = folder / f"{safe_name}.csv"
        row_count = export_sheet_to_csv(destination, ws.title)
        sheet_exports.append((ws.title, destination.name, row_count))

    annuals_path = folder / ANNUALS_EXPORT
    with annuals_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=SELF_SEEDING_MASTER_FIELDS)
        writer.writeheader()
        for row in SELF_SEEDING_MASTER:
            writer.writerow(row)

    inventory_readme_lines = [
        "# Inventory assets",
        "",
        "This folder contains the local zone-7 plant inventories that will be useful when building the ecological-agriculture Codex skill.",
        "",
        "## Files",
        "",
        f"- `{WORKBOOK_PATH.name}`: copied from the workspace as the working perennial inventory workbook.",
    ]
    for sheet_name, filename, row_count in sheet_exports:
        inventory_readme_lines.append(
            f"- `{filename}`: CSV export of workbook sheet `{sheet_name}` with {row_count} data rows."
        )
    inventory_readme_lines.extend(
        [
            f"- `{annuals_path.name}`: curated self-seeding annual and short-lived edible list for zone 7.",
            "",
            "## Notes",
            "",
            "- The workbook is preserved as the source-of-truth local perennial reference.",
            "- The annuals-and-short-lived CSV is a skill-friendly companion list that focuses on species known to reseed, volunteer, or persist through self-sowing in zone 7.",
            "- Categories mostly mirror the workbook taxonomy, with `Fruiting vegetables` added because it is warranted for ground cherries and tomatillos.",
        ]
    )
    (folder / "README.md").write_text("\n".join(inventory_readme_lines) + "\n", encoding="utf-8")


def build_readme(results: list[dict[str, str]]) -> None:
    lines: list[str] = [
        "# Ecological Agriculture Knowledge Base",
        "",
        "This folder was built to support a future Codex skill on ecological agriculture, with Bec Hellouin as the anchor paper and a curated set of peer-reviewed papers or stable link records around small-scale agroecological production, composting, microbes, forest gardens, companion planting, edible wild-plant integration, and region-relevant Mid-Atlantic references.",
        "",
        "## Notes on capture method",
        "",
        "- Most entries below are local PDFs saved into topic folders.",
        "- A smaller set are saved as `.url` shortcuts plus local `.md` notes because direct PDF capture was blocked by publisher anti-bot pages from this environment.",
        "- `catalog.csv` is the machine-readable index for the paper set.",
        "- `Bec_Hellouin_summary.md` is the local study summary.",
        f"- `{TOPICS['inventories_and_local_lists']}` contains the copied perennial workbook, CSV exports of its sheets, and a separate self-seeding annuals-and-short-lived list.",
        f"- `{TOPICS['beneficial_insects_pest_indicators']}` also contains a synthesis note on beneficial plants, beneficial insects, and reading pest pressure as a management signal.",
        "",
        "## Folder map",
        "",
    ]

    for folder in TOPICS.values():
        lines.append(f"- `{folder}`")

    lines.extend(
        [
            "",
            "## Bec Hellouin anchor study",
            "",
            "- Intensive nucleus studied: 1,000 m2, not the whole farm.",
            "- Protected cultivation share: about 42%.",
            "- Recorded themes: hand-scale intensification, intercropping, season extension, ergonomics, labor, and local-market fit.",
            "- Main caution: the intensive core depends on the wider ecological and commercial whole-farm system.",
            "",
            "## Mid-Atlantic priority subset",
            "",
        ]
    )

    for item in [r for r in results if r["mid_atlantic"] == "yes"]:
        lines.append(
            f"- `{TOPICS[item['topic']]}/{item['filename']}` - {item['title']} ({item['year']})"
        )

    lines.extend(
        [
            "",
            "## Inventories and local lists",
            "",
            f"- `{TOPICS['inventories_and_local_lists']}/{WORKBOOK_PATH.name}` - copied perennial workbook from the workspace.",
            f"- `{TOPICS['inventories_and_local_lists']}/{PERENNIAL_SHEET_EXPORT}` - perennial sheet export.",
            f"- `{TOPICS['inventories_and_local_lists']}/{SELF_SEEDING_SHEET_EXPORT}` - workbook self-seeding sheet export.",
            f"- `{TOPICS['inventories_and_local_lists']}/{ANNUALS_EXPORT}` - companion annuals-and-short-lived reseeding list.",
            "",
            "## Catalog by folder",
            "",
        ]
    )

    for topic_key, folder in TOPICS.items():
        lines.append(f"### {folder}")
        lines.append("")
        for item in [r for r in results if r["topic"] == topic_key]:
            lines.append(
                f"- `{item['filename']}` - {item['title']} ({item['year']}). {item['why']}"
            )
        if topic_key == "beneficial_insects_pest_indicators":
            lines.append(
                "- `beneficial_plants_bugs_and_pest_indicators.md` - local synthesis note linking the papers to practical pest-diagnostic and insectary design choices."
            )
        if topic_key == "inventories_and_local_lists":
            lines.extend(
                [
                    f"- `{WORKBOOK_PATH.name}` - local perennial workbook copied into the knowledge base.",
                    f"- `{PERENNIAL_SHEET_EXPORT}` - skill-friendly export of the perennial sheet.",
                    f"- `{SELF_SEEDING_SHEET_EXPORT}` - skill-friendly export of the workbook self-seeding sheet.",
                    f"- `{ANNUALS_EXPORT}` - curated companion list of annual and short-lived self-sowing edibles.",
                ]
            )
        lines.append("")

    (KB_ROOT / "README.md").write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def materialize_resources() -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    for item in RESOURCES:
        folder = KB_ROOT / TOPICS[item["topic"]]
        target = folder / item["filename"]
        if item["kind"] == "copy":
            copy_local(item["source_path"], target)
        elif item["kind"] == "pdf":
            download_pdf(item["download_url"], target)
        elif item["kind"] == "url":
            write_url_shortcut(target, item["source_url"])
            write_url_note(target.with_suffix(".md"), item)
        else:
            raise ValueError(f"Unsupported resource kind: {item['kind']}")
        results.append(item)
    return results


def main() -> None:
    ensure_dirs()
    results = materialize_resources()
    build_bec_hellouin_summary()
    build_beneficial_pest_note()
    build_catalog()
    build_inventory_assets()
    build_readme(results)
    print(f"Knowledge base created at: {KB_ROOT}")
    print(f"Paper/link resource count: {len(results)}")
    print(f"Self-seeding annuals-and-short-lived list count: {len(SELF_SEEDING_MASTER)}")
    for key, folder in TOPICS.items():
        count = sum(1 for item in results if item["topic"] == key)
        print(f"{folder}: {count}")


if __name__ == "__main__":
    main()
