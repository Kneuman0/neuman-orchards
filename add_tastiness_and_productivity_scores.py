from copy import copy
from pathlib import Path
import re

from openpyxl import load_workbook


WORKBOOK = Path(r"G:\My Drive\Home Docs\Garden\Zone_7_Perennial_Subsistence_Crops__comprehensive_list_.xlsx")
TEMP_WORKBOOK = Path.home() / "AppData" / "Local" / "Temp" / "zone7_inventory_score_update.xlsx"
TASTE_SCORE_HEADER = "Tastiness (1-5)"
PRODUCTIVITY_HEADER = "Productivity (1-5)"
HARVEST_HEADER = "Harvest difficulty (1-5)"


def normalize(value):
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


def parse_notes(text):
    text = text or ""
    keys = ["Status", "Tastiness", "Flavor", "Best prep", "Cautions", "Notes"]
    matches = list(re.finditer(r"(Status|Tastiness|Flavor|Best prep|Cautions|Notes):", text))
    parsed = {key: "" for key in keys}
    for idx, match in enumerate(matches):
        key = match.group(1)
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        parsed[key] = text[start:end].strip().strip(". ")
    return parsed


TASTE_SCORE_OVERRIDES = {
    normalize("Asparagus officinalis"): 5,
    normalize("Prunus persica"): 5,
    normalize("Apios americana"): 5,
    normalize("Helianthus tuberosus"): 4,
    normalize("Sium sisarum"): 4,
    normalize("Crambe maritima"): 5,
    normalize("Brassica oleracea"): 5,
    normalize("Allium x proliferum"): 4,
    normalize("Allium fistulosum"): 4,
    normalize("Allium schoenoprasum"): 5,
    normalize("Levisticum officinale"): 5,
    normalize("Origanum vulgare"): 5,
    normalize("Thymus vulgaris"): 5,
    normalize("Salvia officinalis"): 4,
    normalize("Salvia rosmarinus"): 5,
    normalize("Mentha spp."): 5,
    normalize("Artemisia dracunculus"): 5,
    normalize("Melissa officinalis"): 4,
    normalize("Foeniculum vulgare"): 4,
    normalize("Passiflora incarnata"): 5,
    normalize("Vitis vinifera"): 5,
    normalize("Vitis labrusca"): 5,
    normalize("Vitis rotundifolia"): 5,
    normalize("Malus domestica"): 5,
    normalize("Pyrus communis"): 5,
    normalize("Pyrus pyrifolia"): 5,
    normalize("Asimina triloba"): 5,
    normalize("Diospyros virginiana"): 5,
    normalize("Diospyros kaki"): 5,
    normalize("Morus spp."): 5,
    normalize("Ziziphus jujuba"): 5,
    normalize("Ficus carica"): 5,
    normalize("Prunus armeniaca"): 5,
    normalize("Prunus domestica"): 5,
    normalize("Prunus salicina"): 5,
    normalize("Prunus cerasus"): 4,
    normalize("Prunus avium"): 5,
    normalize("Vaccinium corymbosum"): 5,
    normalize("Vaccinium virgatum"): 5,
    normalize("Amelanchier spp."): 5,
    normalize("Ribes rubrum"): 4,
    normalize("Ribes nigrum"): 4,
    normalize("Ribes uva-crispa"): 4,
    normalize("Rubus idaeus"): 5,
    normalize("Rubus occidentalis"): 5,
    normalize("Rubus allegheniensis"): 5,
    normalize("Elaeagnus multiflora"): 4,
    normalize("Corylus spp."): 5,
    normalize("Castanea spp."): 5,
    normalize("Juglans nigra"): 4,
    normalize("Juglans regia"): 5,
    normalize("Carya illinoinensis"): 5,
    normalize("Nasturtium officinale"): 5,
    normalize("Oenanthe javanica"): 4,
    normalize("Thinopyrum intermedium"): 2,
    normalize("Tripsacum dactyloides"): 2,
    normalize("Opuntia humifusa"): 4,
    normalize("Opuntia engelmannii"): 4,
    normalize("Houttuynia cordata"): 2,
    normalize("Phytolacca americana"): 1,
    normalize("Gleditsia triacanthos"): 2,
    normalize("Quercus spp."): 2,
    normalize("Nuphar advena"): 1,
}


PRODUCTIVITY_OVERRIDES = {
    normalize("Helianthus tuberosus"): 5,
    normalize("Apios americana"): 4,
    normalize("Sium sisarum"): 3,
    normalize("Stachys affinis"): 3,
    normalize("Dioscorea polystachya"): 2,
    normalize("Claytonia virginica"): 1,
    normalize("Camassia quamash"): 2,
    normalize("Camassia scilloides"): 1,
    normalize("Eutrema japonicum"): 1,
    normalize("Asparagus officinalis"): 5,
    normalize("Rheum x hybridum"): 4,
    normalize("Crambe maritima"): 4,
    normalize("Blitum bonus-henricus"): 3,
    normalize("Rumex acetosa"): 5,
    normalize("Rumex scutatus"): 4,
    normalize("Diplotaxis tenuifolia"): 4,
    normalize("Urtica dioica"): 4,
    normalize("Hosta spp."): 3,
    normalize("Tilia americana"): 2,
    normalize("Brassica oleracea"): 5,
    normalize("Hemerocallis fulva"): 4,
    normalize("Allium x proliferum"): 5,
    normalize("Allium fistulosum"): 5,
    normalize("Allium schoenoprasum"): 4,
    normalize("Allium tuberosum"): 4,
    normalize("Allium ampeloprasum"): 4,
    normalize("Allium tricoccum"): 1,
    normalize("Levisticum officinale"): 5,
    normalize("Origanum vulgare"): 4,
    normalize("Thymus vulgaris"): 4,
    normalize("Salvia officinalis"): 4,
    normalize("Salvia rosmarinus"): 4,
    normalize("Mentha spp."): 5,
    normalize("Melissa officinalis"): 4,
    normalize("Foeniculum vulgare"): 4,
    normalize("Armoracia rusticana"): 4,
    normalize("Camellia sinensis"): 2,
    normalize("Ilex vomitoria"): 3,
    normalize("Gleditsia triacanthos"): 1,
    normalize("Ceanothus americanus"): 1,
    normalize("Caragana arborescens"): 2,
    normalize("Desmanthus illinoensis"): 2,
    normalize("Passiflora incarnata"): 4,
    normalize("Actinidia arguta"): 4,
    normalize("Actinidia kolomikta"): 3,
    normalize("Vitis vinifera"): 5,
    normalize("Vitis labrusca"): 5,
    normalize("Vitis rotundifolia"): 5,
    normalize("Hablitzia tamnoides"): 4,
    normalize("Malus domestica"): 5,
    normalize("Pyrus communis"): 5,
    normalize("Pyrus pyrifolia"): 5,
    normalize("Asimina triloba"): 4,
    normalize("Morus spp."): 5,
    normalize("Prunus persica"): 5,
    normalize("Prunus armeniaca"): 4,
    normalize("Prunus domestica"): 4,
    normalize("Prunus salicina"): 4,
    normalize("Prunus cerasus"): 4,
    normalize("Prunus avium"): 4,
    normalize("Punica granatum"): 4,
    normalize("Vaccinium corymbosum"): 5,
    normalize("Vaccinium virgatum"): 5,
    normalize("Sambucus canadensis"): 3,
    normalize("Aronia melanocarpa"): 2,
    normalize("Amelanchier spp."): 4,
    normalize("Rubus idaeus"): 5,
    normalize("Rubus occidentalis"): 4,
    normalize("Rubus allegheniensis"): 5,
    normalize("Elaeagnus multiflora"): 4,
    normalize("Hippophae rhamnoides"): 3,
    normalize("Elaeagnus umbellata"): 4,
    normalize("Rosa spp."): 2,
    normalize("Lonicera caerulea"): 4,
    normalize("Corylus spp."): 4,
    normalize("Castanea spp."): 4,
    normalize("Juglans nigra"): 2,
    normalize("Juglans regia"): 2,
    normalize("Carya spp."): 2,
    normalize("Carya illinoinensis"): 2,
    normalize("Quercus spp."): 2,
    normalize("Sagittaria latifolia"): 3,
    normalize("Typha latifolia"): 3,
    normalize("Nasturtium officinale"): 5,
    normalize("Oenanthe javanica"): 4,
    normalize("Thinopyrum intermedium"): 2,
    normalize("Tripsacum dactyloides"): 2,
    normalize("Helianthus maximiliani"): 3,
    normalize("Opuntia humifusa"): 4,
    normalize("Opuntia engelmannii"): 4,
    normalize("Yucca filamentosa"): 2,
    normalize("Agave parryi"): 1,
    normalize("Phytolacca americana"): 1,
}


def phrase_score(text):
    text = (text or "").lower()
    if any(term in text for term in ["barely edible", "survival crop", "not generally regarded as tasty"]):
        return 1
    if any(term in text for term in ["mainly considered useful rather than notably tasty", "usable more than craveable"]):
        return 2
    if any(term in text for term in ["opinions are mixed", "mixed fresh", "pleasant but mild", "considered tasty mainly after the right preparation", "very tart but often considered tasty when sweetened", "can be very good, but only after significant processing"]):
        return 3
    if "generally regarded as tasty" in text or "pleasant in seasoning, tea, or syrup use" in text:
        return 4
    if "widely regarded as tasty" in text:
        return 5
    return None


def note_phrase_for_score(score, category):
    if score == 1:
        return "mainly a survival food or only marginally palatable"
    if score == 2:
        return "mainly considered useful rather than notably tasty"
    if score == 3:
        return "opinions are mixed or it is only good in certain preparations"
    if score == 4:
        if category in {"Culinary herbs", "Tea, spice & sweetener herbs", "Alliums"}:
            return "generally regarded as tasty in culinary use"
        return "generally regarded as tasty"
    if score == 5:
        if category in {"Culinary herbs", "Tea, spice & sweetener herbs", "Alliums", "Leafy greens", "Shoots, stalks & flower vegetables"}:
            return "widely regarded as highly tasty in its culinary role"
        return "widely regarded as highly tasty"
    return ""


def base_taste_score(category, parsed):
    phrase = phrase_score(parsed["Tastiness"])
    if phrase is not None:
        return phrase

    category_defaults = {
        "Roots, tubers & rhizomes": 3,
        "Leafy greens": 4,
        "Shoots, stalks & flower vegetables": 4,
        "Alliums": 4,
        "Culinary herbs": 4,
        "Tea, spice & sweetener herbs": 3,
        "Legumes & protein crops": 2,
        "Vines & climbers": 4,
        "Tree fruits": 4,
        "Shrub, cane & bramble fruits": 4,
        "Nuts & oil crops": 4,
        "Aquatic & wetland edibles": 3,
        "Grains, seeds & pseudocereals": 2,
        "Succulent/xeric edibles": 3,
    }
    return category_defaults.get(category, 3)


def base_productivity_score(category, harvest_window, edible_parts, cautions):
    score = {
        "Roots, tubers & rhizomes": 3,
        "Leafy greens": 4,
        "Shoots, stalks & flower vegetables": 4,
        "Alliums": 4,
        "Culinary herbs": 4,
        "Tea, spice & sweetener herbs": 2,
        "Legumes & protein crops": 2,
        "Vines & climbers": 3,
        "Tree fruits": 4,
        "Shrub, cane & bramble fruits": 4,
        "Nuts & oil crops": 2,
        "Aquatic & wetland edibles": 2,
        "Grains, seeds & pseudocereals": 2,
        "Succulent/xeric edibles": 2,
    }.get(category, 3)

    parts = (edible_parts or "").lower()
    window = (harvest_window or "").lower()
    caution_text = (cautions or "").lower()

    if ";" in parts:
        score += 1
    if any(term in window for term in ["spring - fall", "year-round", "spring - summer", "summer - fall"]):
        score += 1
    if any(term in window for term in ["early spring", "late spring", "late fall", "fall - winter", "fall"]):
        score -= 0 if category in {"Tree fruits", "Shrub, cane & bramble fruits", "Nuts & oil crops"} else 1
    if any(term in caution_text for term in ["toxic", "slow growing", "overharvest", "leach", "processing", "expert"]):
        score -= 1

    return max(1, min(5, int(score)))


def build_notes(parsed, tastiness_phrase):
    parts = []
    if parsed["Status"]:
        parts.append(f"Status: {parsed['Status']}.")
    if tastiness_phrase:
        parts.append(f"Tastiness: {tastiness_phrase}.")
    if parsed["Flavor"]:
        parts.append(f"Flavor: {parsed['Flavor']}.")
    if parsed["Best prep"]:
        parts.append(f"Best prep: {parsed['Best prep']}.")
    if parsed["Cautions"]:
        parts.append(f"Cautions: {parsed['Cautions']}.")
    if parsed["Notes"]:
        parts.append(f"Notes: {parsed['Notes']}.")
    return " ".join(parts)


def ensure_column(ws, headers, after_header, new_header):
    if new_header in headers:
        return
    insert_at = headers.index(after_header) + 2
    ws.insert_cols(insert_at, 1)
    ws.cell(1, insert_at).value = new_header
    template = ws.cell(1, insert_at - 1)
    ws.cell(1, insert_at).style = template.style
    ws.cell(1, insert_at).font = copy(template.font)
    ws.cell(1, insert_at).fill = copy(template.fill)
    ws.cell(1, insert_at).border = copy(template.border)
    ws.cell(1, insert_at).alignment = copy(template.alignment)
    ws.column_dimensions[ws.cell(1, insert_at).column_letter].width = 18


def main():
    wb = load_workbook(WORKBOOK)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

    ensure_column(ws, headers, HARVEST_HEADER, TASTE_SCORE_HEADER)
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    ensure_column(ws, headers, TASTE_SCORE_HEADER, PRODUCTIVITY_HEADER)
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    idx = {header: pos + 1 for pos, header in enumerate(headers)}

    for row in range(2, ws.max_row + 1):
        scientific_name = ws.cell(row, idx["Scientific name"]).value or ""
        common_name = ws.cell(row, idx["Common name"]).value or ""
        category = ws.cell(row, idx["Category"]).value or ""
        harvest_window = ws.cell(row, idx["Harvest window"]).value or ""
        edible_parts = ws.cell(row, idx["Edible parts / uses"]).value or ""
        notes = ws.cell(row, idx["Notes"]).value or ""
        parsed = parse_notes(notes)
        key = normalize(scientific_name) or normalize(common_name)

        taste_score = TASTE_SCORE_OVERRIDES.get(key, base_taste_score(category, parsed))
        productivity_score = PRODUCTIVITY_OVERRIDES.get(
            key,
            base_productivity_score(category, harvest_window, edible_parts, parsed["Cautions"]),
        )
        tastiness_phrase = note_phrase_for_score(taste_score, category)

        ws.cell(row, idx[TASTE_SCORE_HEADER]).value = taste_score
        ws.cell(row, idx[PRODUCTIVITY_HEADER]).value = productivity_score
        ws.cell(row, idx["Notes"]).value = build_notes(parsed, tastiness_phrase)

    wb.save(TEMP_WORKBOOK)
    print(f"Saved staged score update to {TEMP_WORKBOOK}")
    print(f"Rows processed: {ws.max_row - 1}")


if __name__ == "__main__":
    main()
