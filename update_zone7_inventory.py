import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


WORKBOOK = Path(r"G:\My Drive\Home Docs\Garden\Zone_7_Perennial_Subsistence_Crops__comprehensive_list_.xlsx")

HEADERS = [
    "Category",
    "Common name",
    "Scientific name",
    "Growth form",
    "Will you grow it?",
    "Hardiness range",
    "Zone 7 fit",
    "Sun",
    "Water",
    "Harvest window",
    "Edible parts / uses",
    "Notes",
    "Propagation",
    "Nitrogen fixer",
]

CATEGORY_ORDER = [
    "Roots, tubers & rhizomes",
    "Leafy greens",
    "Shoots, stalks & flower vegetables",
    "Alliums",
    "Culinary herbs",
    "Tea, spice & sweetener herbs",
    "Legumes & protein crops",
    "Vines & climbers",
    "Tree fruits",
    "Shrub, cane & bramble fruits",
    "Nuts & oil crops",
    "Aquatic & wetland edibles",
    "Grains, seeds & pseudocereals",
    "Succulent/xeric edibles",
]


def record(
    category,
    common_name,
    scientific_name,
    growth_form,
    hardiness_range,
    zone_7_fit,
    sun,
    water,
    harvest_window,
    edible_parts_uses,
    status,
    taste,
    propagation,
    nitrogen_fixer,
    cautions="",
    extra="",
):
    notes = [f"Status: {status}.", f"Taste: {taste}."]
    if cautions:
        notes.append(f"Cautions: {cautions}.")
    if extra:
        notes.append(f"Notes: {extra}.")
    return {
        "Category": category,
        "Common name": common_name,
        "Scientific name": scientific_name,
        "Growth form": growth_form,
        "Will you grow it?": "",
        "Hardiness range": hardiness_range,
        "Zone 7 fit": zone_7_fit,
        "Sun": sun,
        "Water": water,
        "Harvest window": harvest_window,
        "Edible parts / uses": edible_parts_uses,
        "Notes": " ".join(notes),
        "Propagation": propagation,
        "Nitrogen fixer": nitrogen_fixer,
    }


RECORDS = [
    record("Roots, tubers & rhizomes", "Jerusalem artichoke / sunchoke", "Helianthus tuberosus", "Herbaceous perennial", "3-9", "Hardy", "Full sun", "Low - moderate", "Nov - Mar", "Tubers", "cultivated and wild naturalized", "nutty and sweeter after frost", "Dividing tubers", "No", "spreads aggressively", "Very productive staple tuber crop."),
    record("Roots, tubers & rhizomes", "Groundnut / American hopniss", "Apios americana", "Herbaceous perennial", "3-9", "Hardy", "Sun - part shade", "Moist", "Oct - Nov", "Tubers", "wild native and cultivated", "nutty, potato-bean flavor", "Tuber pieces; seed", "Yes", "", "Climbing legume with edible tubers and seeds."),
    record("Roots, tubers & rhizomes", "Skirret", "Sium sisarum", "Herbaceous perennial", "5-9", "Hardy", "Sun - part shade", "Even moisture", "Nov - Mar", "Roots", "cultivated", "sweet, parsnip-like", "Division; seed", "No", "", "Clump-forming root crop."),
    record("Roots, tubers & rhizomes", "Chinese artichoke / crosne", "Stachys affinis", "Herbaceous perennial", "5-9", "Hardy", "Sun - part shade", "Moderate", "Nov - Jan", "Tubers", "cultivated", "crisp, mild, nutty", "Tuberlets", "No", "can spread slowly", ""),
    record("Roots, tubers & rhizomes", "Chinese yam / cinnamon vine", "Dioscorea polystachya", "Vine", "5-9", "Hardy", "Sun - part shade", "Low - moderate", "Oct - Nov", "Tubers; bulbils", "cultivated and wild naturalized", "starchy and mild", "Bulbils; tuber pieces", "No", "can be invasive; deep tubers are hard to dig", ""),
    record("Roots, tubers & rhizomes", "Spring beauty", "Claytonia virginica", "Herbaceous perennial", "3-8", "Wild in zone 7", "Part shade", "Moderate", "Early spring", "Corms; leaves", "wild native", "mild, slightly sweet", "Seed; division", "No", "small corms make it a minor forage food", ""),
    record("Roots, tubers & rhizomes", "Common camas", "Camassia quamash", "Herbaceous perennial", "4-8", "Hardy", "Sun - part shade", "Moderate", "Late spring - early summer", "Bulbs", "wild native and cultivated", "sweet after long cooking", "Offsets; seed", "No", "must not be confused with toxic death camas", "Traditional cooked bulb crop."),
    record("Roots, tubers & rhizomes", "Atlantic camas", "Camassia scilloides", "Herbaceous perennial", "4-8", "Wild in zone 7", "Sun - part shade", "Moderate", "Late spring", "Bulbs", "wild native", "sweet after cooking", "Seed; offsets", "No", "must not be confused with toxic death camas", ""),
    record("Roots, tubers & rhizomes", "Prairie turnip", "Pediomelum esculentum", "Herbaceous perennial", "3-8", "Hardy", "Full sun", "Low", "Summer - fall", "Roots", "wild native and cultivated", "mild, nutty, starchy", "Seed", "Yes", "", "Dryland indigenous food plant."),
    record("Roots, tubers & rhizomes", "Wasabi", "Eutrema japonicum", "Herbaceous perennial", "6-9", "Hardy with protection", "Part shade", "Constant moisture", "Year-round leaves; roots after 18-24 months", "Rhizomes; leaves", "cultivated", "hot, green, mustard-like", "Division", "No", "needs cool, moist conditions to thrive", ""),
    record("Roots, tubers & rhizomes", "Tiger lily", "Lilium lancifolium", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Moderate", "Fall", "Bulbs; flowers", "cultivated and wild naturalized", "starchy bulbs; mild flowers", "Bulb offsets", "No", "use true tiger lily; lily identity matters", ""),
    record("Roots, tubers & rhizomes", "Violet wood-sorrel", "Oxalis violacea", "Herbaceous perennial", "5-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Spring", "Leaves; flowers; bulbs", "wild native", "tart and lemony", "Bulblets; seed", "No", "contains oxalates", ""),
    record("Leafy greens", "Good King Henry", "Blitum bonus-henricus", "Herbaceous perennial", "4-8", "Hardy", "Sun - part shade", "Moderate", "Spring - summer", "Leaves; shoots", "cultivated", "earthy, spinach-like", "Seed; division", "No", "", "Slow to establish but long lived."),
    record("Leafy greens", "Garden sorrel", "Rumex acetosa", "Herbaceous perennial", "3-9", "Hardy", "Sun - part shade", "Moderate", "Spring - fall", "Leaves", "cultivated and wild naturalized", "bright, lemony", "Division; seed", "No", "contains oxalates", ""),
    record("Leafy greens", "French sorrel", "Rumex scutatus", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Low - moderate", "Spring - fall", "Leaves", "cultivated", "mild, lemony", "Division; seed", "No", "contains oxalates", ""),
    record("Leafy greens", "Perennial arugula", "Diplotaxis tenuifolia", "Herbaceous perennial", "4-9", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves", "cultivated and wild naturalized", "peppery", "Seed; division", "No", "can self-seed", ""),
    record("Leafy greens", "Turkish rocket", "Bunias orientalis", "Herbaceous perennial", "4-8", "Hardy", "Full sun", "Low - moderate", "Spring", "Leaves; shoots", "cultivated and wild naturalized", "mustardy and broccoli-like", "Seed; division", "No", "vigorous and potentially weedy", ""),
    record("Leafy greens", "Stinging nettle", "Urtica dioica", "Herbaceous perennial", "3-10", "Wild in zone 7", "Sun - part shade", "Moderate - moist", "Spring - early summer", "Young leaves", "wild native and cultivated", "deep green, spinach-like", "Division; seed", "No", "stings raw; cook or dry before handling heavily", ""),
    record("Leafy greens", "Wood nettle", "Laportea canadensis", "Herbaceous perennial", "3-8", "Wild in zone 7", "Part shade - shade", "Moist", "Spring - early summer", "Young leaves", "wild native", "tender, nettle-like", "Seed; division", "No", "stings raw", ""),
    record("Leafy greens", "Dandelion", "Taraxacum officinale", "Herbaceous perennial", "3-10", "Wild in zone 7", "Sun - part shade", "Low - moderate", "Spring - fall", "Leaves; roots; flowers", "wild naturalized and cultivated", "pleasantly bitter", "Seed", "No", "", "Common edible volunteer."),
    record("Leafy greens", "Chicory", "Cichorium intybus", "Herbaceous perennial", "3-9", "Wild in zone 7", "Full sun", "Low - moderate", "Spring - fall", "Leaves; roots", "wild naturalized and cultivated", "bitter, nutty when cooked", "Seed", "No", "", ""),
    record("Leafy greens", "Common blue violet", "Viola sororia", "Herbaceous perennial", "3-8", "Wild in zone 7", "Part shade - shade", "Moderate", "Spring", "Leaves; flowers", "wild native and cultivated", "mild leaves; sweet floral blossoms", "Division; seed", "No", "", ""),
    record("Leafy greens", "Broadleaf plantain", "Plantago major", "Herbaceous perennial", "3-9", "Wild in zone 7", "Sun - part shade", "Low - moderate", "Spring - summer", "Young leaves; seeds", "wild naturalized", "mild to earthy", "Seed", "No", "older leaves become fibrous", ""),
    record("Leafy greens", "Patience dock", "Rumex patientia", "Herbaceous perennial", "4-8", "Hardy", "Sun - part shade", "Moderate", "Spring - fall", "Leaves", "cultivated", "mildly tart", "Division; seed", "No", "contains oxalates", ""),
    record("Leafy greens", "Sea beet", "Beta vulgaris subsp. maritima", "Herbaceous perennial", "5-9", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves", "wild native to coasts and cultivated", "salty, earthy", "Seed", "No", "", "Wild beet ancestor."),
    record("Leafy greens", "Oyster leaf", "Mertensia maritima", "Herbaceous perennial", "3-8", "Hardy", "Full sun - part shade", "Moderate", "Spring - summer", "Leaves", "cultivated", "fresh oyster note", "Seed; division", "No", "prefers cool conditions", ""),
    record("Leafy greens", "Hosta", "Hosta spp.", "Herbaceous perennial", "3-9", "Hardy", "Part shade - shade", "Moderate", "Early spring", "Shoots; leaves; buds", "cultivated", "mild, green-bean-like", "Division", "No", "", "Harvest lightly from mature clumps."),
    record("Leafy greens", "American basswood", "Tilia americana", "Tree", "3-8", "Wild in zone 7", "Sun - part shade", "Moderate", "Spring", "Young leaves", "wild native and cultivated", "soft, mild, lettuce-like", "Seed; cuttings", "No", "", "Tender new leaves are the main edible part."),
    record("Shoots, stalks & flower vegetables", "Asparagus", "Asparagus officinalis", "Herbaceous perennial", "3-9", "Hardy", "Full sun", "Moderate", "Spring", "Shoots", "cultivated and wild naturalized", "sweet, grassy", "Crowns; seed", "No", "", ""),
    record("Shoots, stalks & flower vegetables", "Rhubarb", "Rheum x hybridum", "Herbaceous perennial", "3-8", "Hardy", "Full sun", "Moderate", "Spring - early summer", "Leaf stalks", "cultivated", "strongly tart", "Division", "No", "leaf blades are toxic", ""),
    record("Shoots, stalks & flower vegetables", "Sea kale", "Crambe maritima", "Herbaceous perennial", "5-9", "Hardy", "Full sun", "Well-drained", "Spring - summer", "Shoots; leaves; flower buds", "cultivated", "cabbage-like and mild", "Root cuttings; seed", "No", "", ""),
    record("Shoots, stalks & flower vegetables", "Ostrich fern", "Matteuccia struthiopteris", "Herbaceous perennial", "3-7", "Hardy", "Part shade - shade", "Moist", "Early spring", "Fiddleheads", "wild native and cultivated", "green, asparagus-like", "Division", "No", "harvest lightly and only from mature crowns", ""),
    record("Shoots, stalks & flower vegetables", "Daylily", "Hemerocallis fulva", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Low - moderate", "Spring - summer", "Shoots; buds; flowers", "cultivated and wild naturalized", "mild shoots; sweet floral buds", "Division", "No", "some people are sensitive; use known edible daylily types", ""),
    record("Shoots, stalks & flower vegetables", "Perennial kale / broccoli types", "Brassica oleracea", "Herbaceous perennial", "6-9", "Hardy with protection", "Full sun", "Moderate", "Spring - fall", "Leaves; shoots; flower buds", "cultivated", "cabbage-like to broccoli-like", "Cuttings; seed", "No", "best with mulch or protection in colder parts of zone 7", "Covers Daubenton and other perennial brassica forms."),
    record("Shoots, stalks & flower vegetables", "Udo", "Aralia cordata", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Moderate", "Spring", "Blanched shoots", "cultivated", "celery-like, aromatic", "Division", "No", "", "Large Japanese mountain vegetable."),
    record("Shoots, stalks & flower vegetables", "Hops", "Humulus lupulus", "Vine", "3-8", "Hardy", "Full sun", "Moderate", "Spring shoots; late summer cones", "Young shoots", "cultivated and wild naturalized", "asparagus-like shoots; bitter cones", "Rhizomes; cuttings", "No", "vigorous twining vine", ""),
    record("Shoots, stalks & flower vegetables", "Yellow groove bamboo", "Phyllostachys aureosulcata", "Grass-like", "5-10", "Hardy", "Full sun", "Moderate", "Spring", "Young shoots", "cultivated", "mild, bamboo-like", "Rhizome division", "No", "running bamboo can spread aggressively", ""),
    record("Shoots, stalks & flower vegetables", "Common milkweed", "Asclepias syriaca", "Herbaceous perennial", "3-9", "Wild in zone 7", "Full sun", "Low - moderate", "Spring - summer", "Young shoots; buds; very young pods", "wild native", "green-bean-like when properly cooked", "Seed; rhizomes", "No", "requires correct stage and repeated boiling for many preparations; do not confuse with other milkweeds", ""),
    record("Shoots, stalks & flower vegetables", "Eastern redbud", "Cercis canadensis", "Tree", "4-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Spring", "Flowers; young pods", "wild native and cultivated", "sweet-tart flowers", "Seed; cuttings", "Yes", "", ""),
    record("Shoots, stalks & flower vegetables", "American pokeweed", "Phytolacca americana", "Herbaceous perennial", "4-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Very early spring", "Very young shoots", "wild native", "mild after proper boiling", "Seed", "No", "all mature parts are toxic; only expert, traditional preparation is acceptable", ""),
    record("Alliums", "Egyptian walking onion", "Allium x proliferum", "Herbaceous perennial", "3-9", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Greens; bulbs; topsets", "cultivated", "strong onion flavor", "Bulb division; topsets", "No", "", "Self-propagating perennial onion."),
    record("Alliums", "Welsh onion", "Allium fistulosum", "Herbaceous perennial", "4-9", "Hardy", "Full sun", "Moderate", "Spring - fall", "Leaves; stems", "cultivated", "fresh onion flavor", "Division; seed", "No", "", ""),
    record("Alliums", "Chives", "Allium schoenoprasum", "Herbaceous perennial", "3-9", "Hardy", "Sun - part shade", "Low - moderate", "Spring - fall", "Leaves; flowers", "cultivated and wild naturalized", "mild onion flavor", "Division; seed", "No", "", ""),
    record("Alliums", "Garlic chives", "Allium tuberosum", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Low - moderate", "Spring - fall", "Leaves; flower buds", "cultivated and wild naturalized", "garlicky", "Division; seed", "No", "can self-seed", ""),
    record("Alliums", "Elephant garlic / perennial leek", "Allium ampeloprasum", "Herbaceous perennial", "5-9", "Hardy", "Full sun", "Moderate", "Spring greens; summer scapes; fall bulbs", "Leaves; scapes; bulbs", "cultivated", "mild garlic-leek flavor", "Cloves; offsets", "No", "", ""),
    record("Alliums", "Ramps", "Allium tricoccum", "Herbaceous perennial", "3-7", "Wild in zone 7", "Part shade - shade", "Moist", "Early spring", "Leaves; bulbs", "wild native and cultivated", "strong garlicky-onion flavor", "Seed; bulb division", "No", "slow growing; avoid overharvest from wild patches", ""),
    record("Alliums", "Meadow garlic", "Allium canadense", "Herbaceous perennial", "4-8", "Wild in zone 7", "Full sun", "Low - moderate", "Spring", "Leaves; bulbs; bulbils", "wild native", "onion-garlic flavor", "Bulbils; seed", "No", "", ""),
    record("Alliums", "Nodding onion", "Allium cernuum", "Herbaceous perennial", "4-8", "Wild in zone 7", "Full sun", "Low - moderate", "Spring - summer", "Leaves; bulbs; flowers", "wild native and cultivated", "mild onion flavor", "Seed; division", "No", "", ""),
    record("Alliums", "Crow garlic", "Allium vineale", "Herbaceous perennial", "4-8", "Wild in zone 7", "Full sun", "Low - moderate", "Spring", "Leaves; bulbs", "wild naturalized", "strong garlic flavor", "Bulbils; division", "No", "can be weedy in lawns and beds", ""),
    record("Alliums", "German garlic", "Allium senescens", "Herbaceous perennial", "4-8", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves; flowers", "cultivated", "mild garlic-onion flavor", "Division; seed", "No", "", ""),
    record("Culinary herbs", "Lovage", "Levisticum officinale", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Moderate", "Spring - fall", "Leaves; stems; seeds", "cultivated", "strong celery flavor", "Division; seed", "No", "", ""),
    record("Culinary herbs", "Oregano", "Origanum vulgare", "Herbaceous perennial", "4-9", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves", "cultivated and wild naturalized", "warm, pungent, savory", "Division; cuttings", "No", "", ""),
    record("Culinary herbs", "Thyme", "Thymus vulgaris", "Subshrub", "5-9", "Hardy", "Full sun", "Low", "Spring - fall", "Leaves", "cultivated", "warm, resinous, savory", "Cuttings; division", "No", "needs drainage", ""),
    record("Culinary herbs", "Sage", "Salvia officinalis", "Subshrub", "4-8", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves", "cultivated", "savory, piney, camphor-like", "Cuttings; layering", "No", "benefits from renewal pruning", ""),
    record("Culinary herbs", "Rosemary", "Salvia rosmarinus", "Subshrub", "7-10", "Hardy with protection", "Full sun", "Low", "Year-round", "Leaves", "cultivated", "resinous, piney", "Cuttings", "No", "hardiness depends on site and cultivar", ""),
    record("Culinary herbs", "Mint", "Mentha spp.", "Herbaceous perennial", "3-9", "Hardy", "Sun - part shade", "Moderate", "Spring - fall", "Leaves", "cultivated and wild naturalized", "cool and aromatic", "Division; runners", "No", "spreads aggressively", ""),
    record("Culinary herbs", "French tarragon", "Artemisia dracunculus", "Herbaceous perennial", "4-8", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves", "cultivated", "sweet anise flavor", "Division; cuttings", "No", "", ""),
    record("Culinary herbs", "Lemon balm", "Melissa officinalis", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Moderate", "Spring - fall", "Leaves", "cultivated and wild naturalized", "bright lemon scent", "Division; seed", "No", "can self-seed", ""),
    record("Culinary herbs", "Hyssop", "Hyssopus officinalis", "Subshrub", "4-9", "Hardy", "Full sun", "Low - moderate", "Summer", "Leaves; flowers", "cultivated", "minty, camphor-like", "Cuttings; seed", "No", "", ""),
    record("Culinary herbs", "Anise hyssop", "Agastache foeniculum", "Herbaceous perennial", "4-8", "Hardy", "Full sun", "Moderate", "Summer - fall", "Leaves; flowers", "wild native and cultivated", "sweet licorice note", "Seed; division", "No", "", ""),
    record("Culinary herbs", "Wild bergamot", "Monarda fistulosa", "Herbaceous perennial", "3-9", "Wild in zone 7", "Full sun", "Moderate", "Summer", "Leaves; flowers", "wild native and cultivated", "oregano-mint profile", "Division; seed", "No", "", ""),
    record("Culinary herbs", "Bee balm", "Monarda didyma", "Herbaceous perennial", "4-9", "Hardy", "Sun - part shade", "Moderate", "Summer", "Leaves; flowers", "wild native and cultivated", "minty, citrusy", "Division; seed", "No", "", ""),
    record("Culinary herbs", "Winter savory", "Satureja montana", "Subshrub", "5-9", "Hardy", "Full sun", "Low", "Spring - fall", "Leaves", "cultivated", "peppery, savory", "Cuttings; division", "No", "needs good drainage", ""),
    record("Culinary herbs", "Sweet cicely", "Myrrhis odorata", "Herbaceous perennial", "3-7", "Hardy", "Part shade", "Moderate", "Spring - summer", "Leaves; seeds; roots", "cultivated", "sweet anise flavor", "Seed; division", "No", "", ""),
    record("Culinary herbs", "Fennel", "Foeniculum vulgare", "Herbaceous perennial", "6-9", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves; stems; seeds", "cultivated and wild naturalized", "sweet anise flavor", "Seed", "No", "can self-seed", ""),
    record("Culinary herbs", "Salad burnet", "Sanguisorba minor", "Herbaceous perennial", "4-8", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves", "cultivated and wild naturalized", "cool cucumber note", "Seed; division", "No", "", ""),
    record("Culinary herbs", "Fish mint", "Houttuynia cordata", "Herbaceous perennial", "5-10", "Hardy", "Sun - part shade", "Moist", "Spring - fall", "Leaves; shoots", "cultivated", "citrus-fishy, very distinctive", "Division", "No", "spreads aggressively", ""),
    record("Culinary herbs", "Costmary", "Tanacetum balsamita", "Herbaceous perennial", "4-8", "Hardy", "Full sun", "Moderate", "Spring - summer", "Leaves", "cultivated", "minty and balsamic", "Division", "No", "", ""),
    record("Culinary herbs", "Culinary lavender", "Lavandula angustifolia", "Subshrub", "5-9", "Hardy", "Full sun", "Low", "Summer", "Flowers; leaves", "cultivated", "floral, resinous", "Cuttings", "No", "needs drainage", ""),
    record("Culinary herbs", "Honewort / mitsuba", "Cryptotaenia canadensis", "Herbaceous perennial", "3-8", "Wild in zone 7", "Part shade", "Moderate", "Spring - fall", "Leaves; stems", "wild native and cultivated", "parsley-celery flavor", "Seed; division", "No", "", ""),
    record("Tea, spice & sweetener herbs", "Tea plant", "Camellia sinensis", "Shrub", "7-9", "Hardy with protection", "Part sun", "Moderate", "Spring - fall flushes", "Young leaves", "cultivated", "green, tannic, grassy", "Cuttings", "No", "best with acidic soil and shelter from hard freezes", ""),
    record("Tea, spice & sweetener herbs", "Horseradish", "Armoracia rusticana", "Herbaceous perennial", "3-9", "Hardy", "Full sun", "Moderate", "Fall - winter", "Roots", "cultivated and wild naturalized", "hot and pungent", "Root cuttings", "No", "spreads if roots are left behind", ""),
    record("Tea, spice & sweetener herbs", "Sassafras", "Sassafras albidum", "Tree", "4-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Spring leaves; fall roots", "Leaves; roots", "wild native", "citrusy root-beer note", "Seed; suckers", "No", "root use is controversial due safrole; leaves are the lower-risk culinary part", ""),
    record("Tea, spice & sweetener herbs", "Spicebush", "Lindera benzoin", "Shrub", "4-9", "Wild in zone 7", "Part shade", "Moderate", "Fall", "Berries; leaves", "wild native and cultivated", "allspice-citrus note", "Seed; cuttings", "No", "", ""),
    record("Tea, spice & sweetener herbs", "Wintergreen", "Gaultheria procumbens", "Subshrub", "3-8", "Hardy", "Part shade", "Moist, acidic", "Fall - winter", "Leaves; berries", "wild native and cultivated", "sweet minty wintergreen", "Division; seed", "No", "", ""),
    record("Tea, spice & sweetener herbs", "Mountain mint", "Pycnanthemum virginianum", "Herbaceous perennial", "4-8", "Wild in zone 7", "Full sun", "Moderate", "Summer", "Leaves", "wild native and cultivated", "strong mint", "Division; seed", "No", "vigorous spreader", ""),
    record("Tea, spice & sweetener herbs", "Smooth sumac", "Rhus glabra", "Shrub", "3-9", "Wild in zone 7", "Full sun", "Low", "Late summer - fall", "Fruit clusters", "wild native", "tart, lemony", "Seed; suckers", "No", "avoid poison sumac confusion", ""),
    record("Tea, spice & sweetener herbs", "Yaupon holly", "Ilex vomitoria", "Shrub", "7-10", "Hardy", "Full sun - part shade", "Low - moderate", "Year-round leaves", "Leaves", "wild native and cultivated", "toasty and tea-like", "Cuttings; seed", "No", "", "Native caffeinated tea plant."),
    record("Tea, spice & sweetener herbs", "Sugar maple", "Acer saccharum", "Tree", "3-8", "Hardy", "Sun - part shade", "Moderate", "Late winter sap", "Sap", "wild native and cultivated", "sweet maple syrup", "Seed", "No", "", ""),
    record("Tea, spice & sweetener herbs", "Boxelder", "Acer negundo", "Tree", "3-9", "Wild in zone 7", "Sun - part shade", "Low - moderate", "Late winter sap", "Sap", "wild native and cultivated", "milder maple sweetness", "Seed", "No", "", ""),
    record("Tea, spice & sweetener herbs", "Honey locust", "Gleditsia triacanthos", "Tree", "3-8", "Hardy", "Full sun", "Low", "Fall", "Sweet pod pulp", "wild native and cultivated", "sweet and molasses-like", "Seed", "Yes", "avoid hard seeds and thorns on wild forms", ""),
    record("Tea, spice & sweetener herbs", "Sweet birch", "Betula lenta", "Tree", "3-7", "Wild in zone 7", "Sun - part shade", "Moderate", "Late winter sap", "Sap; twigs", "wild native", "wintergreen-like", "Seed", "No", "", ""),
    record("Tea, spice & sweetener herbs", "New Jersey tea", "Ceanothus americanus", "Subshrub", "4-8", "Wild in zone 7", "Full sun", "Low", "Summer leaves", "Leaves", "wild native", "mild black-tea-like", "Seed; cuttings", "Yes", "", ""),
    record("Legumes & protein crops", "Siberian pea shrub", "Caragana arborescens", "Shrub", "2-7", "Hardy", "Full sun", "Low", "Late summer", "Seeds; young pods", "cultivated", "pea-like", "Seed", "Yes", "", ""),
    record("Legumes & protein crops", "Illinois bundleflower", "Desmanthus illinoensis", "Herbaceous perennial", "5-9", "Hardy", "Full sun", "Low - moderate", "Late summer - fall", "Seeds", "wild native and cultivated", "earthy bean flavor after cooking", "Seed", "Yes", "requires cooking", ""),
    record("Legumes & protein crops", "Alfalfa", "Medicago sativa", "Herbaceous perennial", "3-9", "Hardy", "Full sun", "Low - moderate", "Spring - fall", "Leaves; sprouts", "cultivated and wild naturalized", "green and grassy", "Seed", "Yes", "", "Mostly a minor food or sprouting crop."),
    record("Legumes & protein crops", "Red clover", "Trifolium pratense", "Herbaceous perennial", "3-8", "Hardy", "Sun - part shade", "Moderate", "Spring - fall", "Leaves; flowers", "wild naturalized and cultivated", "mild and beany", "Seed", "Yes", "", ""),
    record("Legumes & protein crops", "White clover", "Trifolium repens", "Herbaceous perennial", "3-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Spring - fall", "Leaves; flowers", "wild naturalized and cultivated", "mild and grassy", "Seed", "Yes", "", ""),
    record("Legumes & protein crops", "Ground plum", "Astragalus crassicarpus", "Herbaceous perennial", "3-8", "Hardy", "Full sun", "Low", "Late spring - summer", "Pods", "wild native", "pea-like", "Seed", "Yes", "astragalus identification matters", ""),
    record("Vines & climbers", "Maypop / passionflower", "Passiflora incarnata", "Vine", "6-10", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit; flowers", "wild native and cultivated", "sweet-tart tropical flavor", "Seed; suckers", "No", "", ""),
    record("Vines & climbers", "Hardy kiwi", "Actinidia arguta", "Vine", "4-8", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit", "cultivated", "sweet, aromatic kiwi flavor", "Cuttings", "No", "", "Needs male and female plants unless self-fertile cultivar is used."),
    record("Vines & climbers", "Arctic kiwi", "Actinidia kolomikta", "Vine", "3-8", "Hardy", "Full sun - part shade", "Moderate", "Late summer", "Fruit", "cultivated", "sweet kiwi flavor", "Cuttings", "No", "", ""),
    record("Vines & climbers", "European grape", "Vitis vinifera", "Vine", "6-9", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit; leaves", "cultivated", "sweet to winey", "Cuttings", "No", "", ""),
    record("Vines & climbers", "Fox grape / labrusca grape", "Vitis labrusca", "Vine", "4-8", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit; leaves", "wild native and cultivated", "sweet with classic grapey aroma", "Cuttings", "No", "", ""),
    record("Vines & climbers", "Muscadine grape", "Vitis rotundifolia", "Vine", "7-10", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit", "wild native and cultivated", "musky, sweet", "Cuttings; layering", "No", "", ""),
    record("Vines & climbers", "Schisandra", "Schisandra chinensis", "Vine", "4-7", "Hardy", "Part shade", "Moderate", "Late summer - fall", "Fruit", "cultivated", "tart, resinous, complex", "Cuttings; layering", "No", "", "Often used more as tea or tonic fruit than fresh eating."),
    record("Vines & climbers", "Chocolate vine", "Akebia quinata", "Vine", "5-8", "Hardy", "Sun - part shade", "Moderate", "Fall", "Fruit pulp; young shoots", "cultivated and wild naturalized", "mildly sweet pulp", "Cuttings; layering", "No", "can be invasive in parts of the eastern U.S.", ""),
    record("Vines & climbers", "Caucasian spinach", "Hablitzia tamnoides", "Vine", "3-8", "Hardy", "Part shade - shade", "Moderate", "Spring - summer", "Leaves; shoots", "cultivated", "mild spinach flavor", "Seed; division", "No", "", ""),
    record("Vines & climbers", "Greenbrier", "Smilax rotundifolia", "Vine", "5-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Spring shoots; fall tubers", "Young shoots; tubers", "wild native", "asparagus-like shoots", "Seed; rhizomes", "No", "thorny vine", ""),
    record("Vines & climbers", "Kudzu", "Pueraria montana var. lobata", "Vine", "5-9", "Wild in zone 7", "Full sun", "Low - moderate", "Summer roots and shoots", "Roots; young shoots; flowers", "wild naturalized", "starchy roots; grape-like jelly from flowers", "Root cuttings; crowns", "Yes", "highly invasive and difficult to control", ""),
    record("Tree fruits", "Apple", "Malus domestica", "Tree", "4-8", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit", "cultivated", "sweet to tart depending on cultivar", "Grafting", "No", "", ""),
    record("Tree fruits", "Crabapple, edible types", "Malus coronaria", "Tree", "4-8", "Wild in zone 7", "Full sun", "Moderate", "Fall", "Fruit", "wild native and cultivated", "sharp to tart-sweet", "Seed; grafting", "No", "many crabapples are mainly culinary rather than fresh eating", ""),
    record("Tree fruits", "Pear", "Pyrus communis", "Tree", "4-8", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit", "cultivated", "sweet and juicy", "Grafting", "No", "", ""),
    record("Tree fruits", "Asian pear", "Pyrus pyrifolia", "Tree", "5-9", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit", "cultivated", "crisp, sweet", "Grafting", "No", "", ""),
    record("Tree fruits", "Quince", "Cydonia oblonga", "Tree", "5-9", "Hardy", "Full sun", "Moderate", "Fall", "Fruit", "cultivated", "fragrant, tart", "Cuttings; grafting", "No", "usually cooked before eating", ""),
    record("Tree fruits", "Medlar", "Mespilus germanica", "Tree", "5-8", "Hardy", "Full sun", "Moderate", "Late fall", "Fruit", "cultivated", "apple-date flavor after bletting", "Grafting", "No", "best eaten bletted", ""),
    record("Tree fruits", "Pawpaw", "Asimina triloba", "Tree", "5-9", "Hardy", "Sun - part shade", "Moderate", "Late summer - fall", "Fruit", "wild native and cultivated", "custardy, banana-mango flavor", "Seed; grafting", "No", "", "Needs cross-pollination."),
    record("Tree fruits", "American persimmon", "Diospyros virginiana", "Tree", "4-9", "Hardy", "Full sun", "Low - moderate", "Fall", "Fruit", "wild native and cultivated", "honey-sweet when fully ripe", "Seed; grafting", "No", "astringent until fully soft", ""),
    record("Tree fruits", "Asian persimmon", "Diospyros kaki", "Tree", "7-10", "Hardy with protection", "Full sun", "Moderate", "Fall", "Fruit", "cultivated", "sweet and rich", "Grafting", "No", "choose hardy cultivars in colder parts of zone 7", ""),
    record("Tree fruits", "Mulberry", "Morus spp.", "Tree", "4-9", "Hardy", "Full sun", "Low - moderate", "Late spring - summer", "Fruit", "wild native and cultivated", "sweet, juicy, winey", "Cuttings; grafting", "No", "birds can strip crops quickly", ""),
    record("Tree fruits", "Jujube", "Ziziphus jujuba", "Tree", "5-9", "Hardy", "Full sun", "Low", "Late summer - fall", "Fruit", "cultivated", "apple-like fresh; date-like dried", "Grafting", "No", "", ""),
    record("Tree fruits", "Fig", "Ficus carica", "Shrub", "6-10", "Hardy with protection", "Full sun", "Low - moderate", "Late summer - fall", "Fruit", "cultivated", "honey-sweet", "Cuttings", "No", "best against warm walls or with winter protection in colder sites", ""),
    record("Tree fruits", "Peach / nectarine", "Prunus persica", "Tree", "5-8", "Hardy", "Full sun", "Moderate", "Summer", "Fruit", "cultivated", "sweet and aromatic", "Grafting", "No", "", ""),
    record("Tree fruits", "Apricot", "Prunus armeniaca", "Tree", "5-8", "Hardy", "Full sun", "Low - moderate", "Summer", "Fruit", "cultivated", "sweet, rich", "Grafting", "No", "late frosts often affect bloom", ""),
    record("Tree fruits", "European plum", "Prunus domestica", "Tree", "4-8", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Fruit", "cultivated", "sweet and rich", "Grafting", "No", "", ""),
    record("Tree fruits", "Japanese plum", "Prunus salicina", "Tree", "5-9", "Hardy", "Full sun", "Moderate", "Summer", "Fruit", "cultivated", "juicy, sweet-tart", "Grafting", "No", "", ""),
    record("Tree fruits", "American plum", "Prunus americana", "Tree", "3-8", "Wild in zone 7", "Full sun", "Moderate", "Late summer", "Fruit", "wild native and cultivated", "sweet-tart", "Seed; suckers", "No", "", "Thicket-forming native plum."),
    record("Tree fruits", "Chickasaw plum", "Prunus angustifolia", "Shrub", "5-9", "Wild in zone 7", "Full sun", "Low - moderate", "Summer", "Fruit", "wild native", "sweet-tart", "Suckers; seed", "No", "", ""),
    record("Tree fruits", "Sour cherry", "Prunus cerasus", "Tree", "4-8", "Hardy", "Full sun", "Moderate", "Early summer", "Fruit", "cultivated", "bright tart cherry flavor", "Grafting", "No", "", ""),
    record("Tree fruits", "Sweet cherry", "Prunus avium", "Tree", "5-8", "Hardy", "Full sun", "Moderate", "Early summer", "Fruit", "cultivated", "sweet and juicy", "Grafting", "No", "", ""),
    record("Tree fruits", "Cornelian cherry dogwood", "Cornus mas", "Tree", "4-8", "Hardy", "Full sun - part shade", "Moderate", "Late summer", "Fruit", "cultivated", "tart cherry-like flavor", "Cuttings; grafting", "No", "", ""),
    record("Tree fruits", "Hackberry", "Celtis occidentalis", "Tree", "3-9", "Wild in zone 7", "Full sun", "Low - moderate", "Fall - winter", "Fruit", "wild native", "sweet, date-like skin over crunchy seed", "Seed", "No", "", ""),
    record("Tree fruits", "Hawthorn", "Crataegus spp.", "Tree", "4-8", "Hardy", "Full sun", "Low - moderate", "Fall", "Fruit", "wild native and cultivated", "mealy apple-like fruit", "Seed; grafting", "No", "thorny; species and fruit quality vary widely", ""),
    record("Tree fruits", "Che / Chinese mulberry", "Cudrania tricuspidata", "Tree", "6-9", "Hardy", "Full sun", "Low - moderate", "Late summer - fall", "Fruit", "cultivated", "sweet and mild", "Cuttings; grafting", "No", "", ""),
    record("Tree fruits", "Pomegranate", "Punica granatum", "Shrub", "7-10", "Hardy with protection", "Full sun", "Low - moderate", "Fall", "Fruit", "cultivated", "sweet-tart", "Cuttings", "No", "hardiness depends heavily on cultivar and site", ""),
    record("Tree fruits", "Trifoliate orange", "Citrus trifoliata", "Shrub", "5-9", "Hardy", "Full sun", "Moderate", "Fall", "Fruit", "cultivated and naturalized", "strongly sour and aromatic", "Seed; grafting", "No", "very seedy and often used more for marmalade or breeding than fresh eating", ""),
    record("Shrub, cane & bramble fruits", "Highbush blueberry", "Vaccinium corymbosum", "Shrub", "3-8", "Hardy", "Full sun", "Moist, acidic", "Summer", "Fruit", "wild native and cultivated", "sweet-tart", "Cuttings", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Rabbiteye blueberry", "Vaccinium virgatum", "Shrub", "6-9", "Hardy", "Full sun", "Moist, acidic", "Summer", "Fruit", "wild native and cultivated", "sweet", "Cuttings", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Cranberry", "Vaccinium macrocarpon", "Trailing shrub", "2-7", "Hardy", "Full sun", "Moist, acidic", "Fall", "Fruit", "wild native and cultivated", "sharp tartness", "Cuttings", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Black huckleberry", "Gaylussacia baccata", "Shrub", "3-8", "Wild in zone 7", "Part shade - sun", "Moist, acidic", "Summer", "Fruit", "wild native", "sweet with crunchy seeds", "Layering; division", "No", "", ""),
    record("Shrub, cane & bramble fruits", "American elderberry", "Sambucus canadensis", "Shrub", "4-9", "Wild in zone 7", "Sun - part shade", "Moderate - moist", "Late summer", "Fruit; flowers", "wild native and cultivated", "dark, winey berries", "Cuttings; suckers", "No", "berries are best cooked; stems and unripe parts are not edible", ""),
    record("Shrub, cane & bramble fruits", "Black chokeberry", "Aronia melanocarpa", "Shrub", "3-8", "Hardy", "Sun - part shade", "Moderate", "Late summer - fall", "Fruit", "wild native and cultivated", "dry, tannic, tart-sweet", "Cuttings; division", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Serviceberry / juneberry", "Amelanchier spp.", "Shrub", "3-9", "Hardy", "Sun - part shade", "Moderate", "Early summer", "Fruit", "wild native and cultivated", "blueberry-almond flavor", "Seed; grafting", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Black currant", "Ribes nigrum", "Shrub", "3-8", "Hardy", "Sun - part shade", "Moderate", "Summer", "Fruit", "cultivated", "deep, musky, tart", "Cuttings", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Red currant", "Ribes rubrum", "Shrub", "3-8", "Hardy", "Sun - part shade", "Moderate", "Summer", "Fruit", "cultivated", "bright tartness", "Cuttings", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Gooseberry", "Ribes uva-crispa", "Shrub", "3-8", "Hardy", "Sun - part shade", "Moderate", "Summer", "Fruit", "cultivated", "tart to sweet depending on ripeness", "Cuttings", "No", "thorny canes", ""),
    record("Shrub, cane & bramble fruits", "Clove currant", "Ribes odoratum", "Shrub", "4-8", "Hardy", "Sun - part shade", "Moderate", "Summer", "Fruit", "wild native and cultivated", "sweet with clove scent", "Cuttings", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Red raspberry", "Rubus idaeus", "Cane", "3-8", "Hardy", "Full sun", "Moderate", "Summer - fall", "Fruit", "cultivated", "sweet-tart", "Division; tip layering", "No", "spreads by suckers", ""),
    record("Shrub, cane & bramble fruits", "Black raspberry", "Rubus occidentalis", "Cane", "3-8", "Wild in zone 7", "Full sun", "Moderate", "Summer", "Fruit", "wild native and cultivated", "rich and winey", "Tip layering", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Blackberry", "Rubus allegheniensis", "Cane", "3-8", "Wild in zone 7", "Full sun", "Moderate", "Summer", "Fruit", "wild native and cultivated", "sweet-tart", "Tip layering; suckers", "No", "thorny on wild forms", ""),
    record("Shrub, cane & bramble fruits", "Goumi", "Elaeagnus multiflora", "Shrub", "4-8", "Hardy", "Full sun", "Low - moderate", "Late spring - early summer", "Fruit", "cultivated", "bright, tart-sweet", "Cuttings", "Yes", "", ""),
    record("Shrub, cane & bramble fruits", "Sea buckthorn", "Hippophae rhamnoides", "Shrub", "3-8", "Hardy", "Full sun", "Low", "Late summer - fall", "Fruit", "cultivated", "sharp, citrusy", "Cuttings; suckers", "Yes", "thorny and usually needs male and female plants", ""),
    record("Shrub, cane & bramble fruits", "Autumn olive", "Elaeagnus umbellata", "Shrub", "4-8", "Wild in zone 7", "Full sun", "Low - moderate", "Fall", "Fruit", "wild naturalized", "sweet-tart", "Seed; cuttings", "Yes", "highly invasive in many eastern states", ""),
    record("Shrub, cane & bramble fruits", "American cranberrybush", "Viburnum trilobum", "Shrub", "2-7", "Hardy", "Sun - part shade", "Moderate", "Fall - winter", "Fruit", "wild native and cultivated", "tart and musky", "Cuttings", "No", "often better cooked than fresh", ""),
    record("Shrub, cane & bramble fruits", "Nannyberry", "Viburnum lentago", "Shrub", "2-8", "Wild in zone 7", "Sun - part shade", "Moderate", "Fall", "Fruit", "wild native", "sweet, raisin-like when ripe", "Seed; suckers", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Rose hips", "Rosa spp.", "Shrub", "3-9", "Hardy", "Full sun", "Moderate", "Fall - winter", "Hips; petals", "wild native and cultivated", "tart and fruity", "Cuttings; suckers", "No", "remove irritating inner hairs when using hips", ""),
    record("Shrub, cane & bramble fruits", "Goji berry", "Lycium barbarum", "Shrub", "5-9", "Hardy", "Full sun", "Low - moderate", "Summer - fall", "Fruit", "cultivated", "sweet-tart", "Cuttings; seed", "No", "can sucker and sprawl", ""),
    record("Shrub, cane & bramble fruits", "Flowering quince", "Chaenomeles speciosa", "Shrub", "4-8", "Hardy", "Full sun", "Moderate", "Fall", "Fruit", "cultivated", "very tart and aromatic", "Cuttings", "No", "usually cooked", ""),
    record("Shrub, cane & bramble fruits", "American beautyberry", "Callicarpa americana", "Shrub", "6-10", "Hardy", "Sun - part shade", "Moderate", "Fall", "Fruit", "wild native and cultivated", "mild and best in jelly", "Cuttings", "No", "", ""),
    record("Shrub, cane & bramble fruits", "Haskap / honeyberry", "Lonicera caerulea", "Shrub", "2-8", "Hardy", "Full sun", "Moderate", "Late spring", "Fruit", "cultivated", "blueberry-meets-raspberry", "Cuttings", "No", "", ""),
    record("Nuts & oil crops", "Hazelnut / filbert", "Corylus spp.", "Shrub", "4-8", "Hardy", "Full sun", "Moderate", "Late summer - fall", "Nuts", "wild native and cultivated", "sweet, rich, nutty", "Layering; grafting", "No", "", "Includes American, European, and hybrid hazels."),
    record("Nuts & oil crops", "Chestnut", "Castanea spp.", "Tree", "4-8", "Hardy", "Full sun", "Moderate", "Fall", "Nuts", "wild native and cultivated", "sweet and starchy", "Grafting; seed", "No", "", "Includes Chinese, American, and hybrid chestnuts."),
    record("Nuts & oil crops", "Black walnut", "Juglans nigra", "Tree", "4-9", "Wild in zone 7", "Full sun", "Moderate", "Fall", "Nuts", "wild native and cultivated", "bold, rich, resinous", "Seed; grafting", "No", "strong juglone effects on nearby plants", ""),
    record("Nuts & oil crops", "English walnut", "Juglans regia", "Tree", "5-9", "Hardy", "Full sun", "Moderate", "Fall", "Nuts", "cultivated", "mild, rich", "Grafting", "No", "", ""),
    record("Nuts & oil crops", "Butternut", "Juglans cinerea", "Tree", "3-7", "Hardy", "Full sun", "Moderate", "Fall", "Nuts", "wild native", "rich and buttery", "Seed", "No", "butternut canker limits many plantings", ""),
    record("Nuts & oil crops", "Hickory", "Carya spp.", "Tree", "4-8", "Hardy", "Full sun", "Moderate", "Fall", "Nuts", "wild native and cultivated", "sweet, rich, pecan-like", "Seed; grafting", "No", "", "Includes shagbark and shellbark types."),
    record("Nuts & oil crops", "Pecan", "Carya illinoinensis", "Tree", "5-9", "Hardy", "Full sun", "Moderate", "Fall", "Nuts", "wild native and cultivated", "sweet, rich", "Grafting", "No", "northern-hardy cultivars are best in cooler zone 7 sites", ""),
    record("Nuts & oil crops", "American beech", "Fagus grandifolia", "Tree", "3-9", "Wild in zone 7", "Sun - part shade", "Moderate", "Fall", "Nuts", "wild native", "small but sweet and rich", "Seed", "No", "", ""),
    record("Nuts & oil crops", "Korean pine", "Pinus koraiensis", "Tree", "3-8", "Hardy", "Full sun", "Moderate", "Fall", "Nuts", "cultivated", "mild pine-nut flavor", "Seed", "No", "", ""),
    record("Nuts & oil crops", "Pinyon pine", "Pinus edulis", "Tree", "5-8", "Hardy", "Full sun", "Low", "Fall", "Nuts", "cultivated", "classic pine-nut flavor", "Seed", "No", "best in drier sites", ""),
    record("Nuts & oil crops", "Tea-oil camellia", "Camellia oleifera", "Shrub", "7-10", "Hardy with protection", "Part sun", "Moderate", "Fall", "Seeds for oil", "cultivated", "oil is mild and buttery", "Cuttings", "No", "best in protected acidic sites", ""),
    record("Nuts & oil crops", "Ginkgo", "Ginkgo biloba", "Tree", "3-9", "Hardy", "Full sun", "Low - moderate", "Fall", "Seeds", "cultivated", "rich, chestnut-like when cooked", "Grafting; seed", "No", "raw seeds are not eaten; female fruit pulp is messy", ""),
    record("Nuts & oil crops", "Sweet-acorn oaks", "Quercus spp.", "Tree", "3-9", "Hardy", "Full sun", "Low - moderate", "Fall", "Acorns", "wild native and cultivated", "nutty and starchy after leaching", "Seed", "No", "tannins require leaching; focus on lower-tannin white-oak-group species", ""),
    record("Aquatic & wetland edibles", "Arrowhead / duck potato", "Sagittaria latifolia", "Aquatic", "3-10", "Hardy", "Full sun", "Aquatic", "Fall", "Corms", "wild native and cultivated", "potato-like", "Division; tubers", "No", "", ""),
    record("Aquatic & wetland edibles", "Broadleaf cattail", "Typha latifolia", "Aquatic", "3-9", "Wild in zone 7", "Full sun", "Aquatic", "Spring shoots; summer pollen; fall rhizomes", "Shoots; pollen; rhizomes", "wild native", "mild shoots; starchy rhizomes", "Rhizome division", "No", "spreads strongly", ""),
    record("Aquatic & wetland edibles", "American lotus", "Nelumbo lutea", "Aquatic", "4-10", "Wild in zone 7", "Full sun", "Aquatic", "Late summer - fall", "Rhizomes; seeds", "wild native and cultivated", "starchy rhizome; nutty seeds", "Rhizome division", "No", "", ""),
    record("Aquatic & wetland edibles", "Sacred lotus", "Nelumbo nucifera", "Aquatic", "4-10", "Hardy", "Full sun", "Aquatic", "Late summer - fall", "Rhizomes; seeds", "cultivated", "starchy rhizome; nutty seeds", "Rhizome division", "No", "", ""),
    record("Aquatic & wetland edibles", "Watercress", "Nasturtium officinale", "Aquatic", "3-9", "Hardy", "Sun - part shade", "Running water or very moist soil", "Cool seasons", "Leaves", "wild naturalized and cultivated", "peppery", "Cuttings; division", "No", "use clean water sources only", ""),
    record("Aquatic & wetland edibles", "Water celery", "Oenanthe javanica", "Aquatic", "6-10", "Hardy", "Sun - part shade", "Wet - aquatic", "Spring - fall", "Leaves; stems", "cultivated", "celery-parsley flavor", "Division", "No", "spreads readily in wet ground", ""),
    record("Aquatic & wetland edibles", "Pickerelweed", "Pontederia cordata", "Aquatic", "3-10", "Wild in zone 7", "Full sun", "Aquatic", "Late summer - fall", "Seeds; young shoots", "wild native", "mild seeds", "Division; seed", "No", "", ""),
    record("Aquatic & wetland edibles", "Yellow pond-lily", "Nuphar advena", "Aquatic", "4-10", "Wild in zone 7", "Full sun", "Aquatic", "Late summer - fall", "Seeds; rhizomes", "wild native", "starchy and nutty when processed", "Rhizome division", "No", "usually requires drying, roasting, or other processing", ""),
    record("Grains, seeds & pseudocereals", "Kernza / intermediate wheatgrass", "Thinopyrum intermedium", "Grass-like", "4-8", "Hardy", "Full sun", "Moderate", "Summer", "Seeds", "cultivated", "mild nutty grain flavor", "Seed; division", "No", "", "Emerging perennial grain crop."),
    record("Grains, seeds & pseudocereals", "Eastern gamagrass", "Tripsacum dactyloides", "Grass-like", "4-9", "Wild in zone 7", "Full sun", "Moderate", "Late summer - fall", "Seeds", "wild native and cultivated", "corn-like grain note", "Seed; division", "No", "", ""),
    record("Grains, seeds & pseudocereals", "Maximilian sunflower", "Helianthus maximiliani", "Herbaceous perennial", "4-9", "Hardy", "Full sun", "Low - moderate", "Fall", "Seeds", "wild native and cultivated", "sunflower-seed flavor", "Seed; division", "No", "tall and spreading", ""),
    record("Grains, seeds & pseudocereals", "Silflower / silphium", "Silphium integrifolium", "Herbaceous perennial", "3-9", "Hardy", "Full sun", "Low - moderate", "Fall", "Seeds", "wild native and cultivated", "oily seed flavor", "Seed", "No", "", "Breeding target for perennial oilseed production."),
    record("Grains, seeds & pseudocereals", "Perennial buckwheat", "Fagopyrum cymosum", "Herbaceous perennial", "6-9", "Hardy", "Full sun - part shade", "Moderate", "Late summer - fall", "Leaves; seeds", "cultivated", "buckwheat-like", "Division; seed", "No", "not as common or proven as annual buckwheat", ""),
    record("Succulent/xeric edibles", "Eastern prickly pear", "Opuntia humifusa", "Succulent", "4-9", "Hardy", "Full sun", "Low", "Summer fruits; spring pads", "Pads; fruit", "wild native and cultivated", "tart fruit; green-bean-like pads", "Pads; seed", "No", "remove spines and glochids carefully", ""),
    record("Succulent/xeric edibles", "Engelmann prickly pear", "Opuntia engelmannii", "Succulent", "6-10", "Hardy", "Full sun", "Low", "Summer", "Pads; fruit", "cultivated", "mild pads; sweet-tart fruit", "Pads", "No", "remove spines and glochids carefully", ""),
    record("Succulent/xeric edibles", "Yucca", "Yucca filamentosa", "Succulent", "4-10", "Hardy", "Full sun", "Low", "Spring - summer", "Flowers; immature fruit", "wild native and cultivated", "mild and slightly sweet flowers", "Offsets", "No", "older mature pods are fibrous", ""),
    record("Succulent/xeric edibles", "Fourwing saltbush", "Atriplex canescens", "Shrub", "3-7", "Hardy", "Full sun", "Low", "Summer leaves; fall seeds", "Leaves; seeds", "wild native and cultivated", "salty and earthy", "Seed", "No", "best used in moderation", ""),
    record("Succulent/xeric edibles", "Mediterranean saltbush", "Atriplex halimus", "Shrub", "7-10", "Hardy", "Full sun", "Low", "Year-round leaves", "Leaves", "cultivated", "salty, spinach-like", "Cuttings; seed", "No", "best in dry, well-drained sites", ""),
    record("Succulent/xeric edibles", "Parry's agave", "Agave parryi", "Succulent", "5-9", "Hardy", "Full sun", "Low", "Any time after maturity", "Heart; flower stalk", "cultivated", "sweet when roasted", "Offsets", "No", "harvesting the heart kills the plant; sharp spines", ""),
]


LEGACY_WILL = {
    "Helianthus tuberosus": "Yes",
    "Apios americana": "Yes",
    "Sium sisarum": "Yes",
    "Matteuccia struthiopteris": "Yes",
    "Hemerocallis fulva": "Yes",
    "Blitum bonus-henricus": "Yes",
    "Levisticum officinale": "Yes",
    "Hosta spp.": "Yes",
    "Allium x proliferum": "Yes",
    "Allium fistulosum": "Yes",
    "Allium schoenoprasum": "Yes",
    "Caragana arborescens": "Yes",
    "Elaeagnus multiflora": "Yes",
    "Actinidia arguta": "Yes",
}


def normalize_key(value):
    if value is None:
        return ""
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"

    widths = {
        "A": 28,
        "B": 30,
        "C": 28,
        "D": 18,
        "E": 16,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 24,
        "K": 20,
        "L": 84,
        "M": 24,
        "N": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def main():
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else WORKBOOK

    wb = load_workbook(WORKBOOK)
    ws = wb[wb.sheetnames[0]]

    existing_will = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        common_name = row[1] if len(row) > 1 else ""
        scientific_name = row[2] if len(row) > 2 else ""
        will_grow = row[3] if len(row) > 3 else ""
        if will_grow:
            existing_will[normalize_key(common_name)] = will_grow
            existing_will[normalize_key(scientific_name)] = will_grow

    scientific_seen = set()
    unique_records = []
    for rec in RECORDS:
        sci_key = normalize_key(rec["Scientific name"])
        if sci_key in scientific_seen:
            continue
        scientific_seen.add(sci_key)

        will_value = existing_will.get(sci_key) or existing_will.get(normalize_key(rec["Common name"])) or LEGACY_WILL.get(rec["Scientific name"], "")
        if will_value:
            rec["Will you grow it?"] = will_value
        unique_records.append(rec)

    unique_records.sort(key=lambda rec: (CATEGORY_ORDER.index(rec["Category"]), rec["Common name"].lower()))

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    if ws.max_column > 0:
        ws.delete_cols(1, ws.max_column)

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, rec in enumerate(unique_records, start=2):
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=rec[header])

    style_sheet(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    wb.save(workbook_path)

    category_counts = {}
    for rec in unique_records:
        category_counts[rec["Category"]] = category_counts.get(rec["Category"], 0) + 1

    print(f"Saved {len(unique_records)} records to {workbook_path.name}")
    for category in CATEGORY_ORDER:
        print(f"{category}: {category_counts.get(category, 0)}")


if __name__ == "__main__":
    main()
