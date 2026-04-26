#!/usr/bin/env python3
"""Generate a garden plot workbook with 2-inch grid cells."""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CELL_SIZE_INCHES = 2
GRID_COLUMN_WIDTH = 26.71
GRID_ROW_HEIGHT_POINTS = 144
HEADER_ROW = 5
GRID_START_ROW = 6
GRID_START_COL = 2
DEFAULT_TITLE = "Garden Plot Layout"
PALETTE = [
    "A9D18E",
    "9DC3E6",
    "FFD966",
    "F4B183",
    "C9C9F7",
    "E2F0D9",
    "F8CBAD",
    "DDEBF7",
    "E4DFEC",
    "FFF2CC",
    "C6E0B4",
    "B4C7E7",
]
STOP_WORDS = {"and", "or", "the", "of", "a"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--spec", required=True, help="Path to the plot spec JSON file.")
    parser.add_argument("--output", required=True, help="Path for the generated workbook.")
    return parser.parse_args()


def require_whole_cells(value: float, field_name: str) -> int:
    cells = float(value) / CELL_SIZE_INCHES
    rounded = round(cells)
    if not math.isclose(cells, rounded, abs_tol=1e-9):
        raise ValueError(
            f"{field_name}={value} is not divisible into {CELL_SIZE_INCHES}-inch cells."
        )
    if rounded <= 0:
        raise ValueError(f"{field_name} must be greater than zero.")
    return int(rounded)


def normalize_hex_color(value: str) -> str:
    cleaned = value.strip().lstrip("#").upper()
    if not re.fullmatch(r"[0-9A-F]{6}", cleaned):
        raise ValueError(f"Invalid hex color: {value}")
    return cleaned


def generate_code(crop_name: str, used_codes: set[str]) -> str:
    words = [w.upper() for w in re.findall(r"[A-Za-z0-9]+", crop_name) if w.lower() not in STOP_WORDS]
    if not words:
        words = ["PLOT"]
    candidates = []
    if len(words) == 1:
        word = words[0]
        candidates.extend([word[:4], word[:3], word[:2]])
    else:
        candidates.append("".join(word[0] for word in words[:4]))
        candidates.append("".join(word[:2] for word in words[:2])[:4])
        candidates.append("".join(word[0] for word in words[:3]))
    for candidate in candidates:
        candidate = candidate[:4]
        if candidate and candidate not in used_codes:
            used_codes.add(candidate)
            return candidate
    base = candidates[0][:3] if candidates and candidates[0] else "PLT"
    for number in range(1, 100):
        candidate = f"{base[:3]}{number}"[:4]
        if candidate not in used_codes:
            used_codes.add(candidate)
            return candidate
    raise ValueError(f"Could not generate a unique crop code for {crop_name!r}.")


def assign_color(crop_name: str, color_map: dict[str, str]) -> str:
    if crop_name in color_map:
        return color_map[crop_name]
    color = PALETTE[len(color_map) % len(PALETTE)]
    color_map[crop_name] = color
    return color


def contrast_font_color(fill_hex: str) -> str:
    red = int(fill_hex[0:2], 16)
    green = int(fill_hex[2:4], 16)
    blue = int(fill_hex[4:6], 16)
    luminance = (0.299 * red) + (0.587 * green) + (0.114 * blue)
    return "000000" if luminance > 170 else "FFFFFF"


def safe_sheet_title(base_name: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[:\\\\/?*\\[\\]]", "-", base_name).strip()
    cleaned = cleaned or "Sheet"
    candidate = cleaned[:31]
    counter = 2
    while candidate in existing:
        suffix = f" {counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing.add(candidate)
    return candidate


def load_spec(spec_path: Path) -> dict:
    raw = json.loads(spec_path.read_text(encoding="utf-8-sig"))
    if not isinstance(raw, dict):
        raise ValueError("Spec root must be a JSON object.")
    if raw.get("cell_size_inches", CELL_SIZE_INCHES) != CELL_SIZE_INCHES:
        raise ValueError("This generator only supports 2-inch cells.")

    title = str(raw.get("title", DEFAULT_TITLE)).strip() or DEFAULT_TITLE
    assumptions = [str(item) for item in raw.get("assumptions", [])]
    beds_raw = raw.get("beds")
    if not isinstance(beds_raw, list) or not beds_raw:
        raise ValueError("Spec must contain a non-empty beds list.")

    code_map: dict[str, str] = {}
    color_map: dict[str, str] = {}
    used_codes: set[str] = set()
    legend: dict[str, dict] = {}
    beds = []

    for bed_raw in beds_raw:
        if not isinstance(bed_raw, dict):
            raise ValueError("Each bed must be an object.")
        bed_name = str(bed_raw.get("name", "")).strip()
        if not bed_name:
            raise ValueError("Each bed must have a name.")
        width_in = float(bed_raw["width_in"])
        length_in = float(bed_raw["length_in"])
        cols = require_whole_cells(width_in, f"{bed_name} width_in")
        rows = require_whole_cells(length_in, f"{bed_name} length_in")
        phases_raw = bed_raw.get("phases")
        if phases_raw is None:
            phases_raw = [
                {
                    "name": str(bed_raw.get("phase_name", "Plan")).strip() or "Plan",
                    "notes": bed_raw.get("phase_notes", ""),
                    "placements": bed_raw.get("placements", []),
                }
            ]
        if not isinstance(phases_raw, list) or not phases_raw:
            raise ValueError(f"{bed_name} must include at least one phase.")

        phases = []
        for phase_raw in phases_raw:
            if not isinstance(phase_raw, dict):
                raise ValueError(f"Each phase in {bed_name} must be an object.")
            phase_name = str(phase_raw.get("name", "")).strip()
            if not phase_name:
                raise ValueError(f"Each phase in {bed_name} must have a name.")
            placements_raw = phase_raw.get("placements", [])
            if not isinstance(placements_raw, list):
                raise ValueError(f"placements for {bed_name} / {phase_name} must be a list.")
            occupied: dict[tuple[int, int], str] = {}
            placements = []
            for placement_raw in placements_raw:
                if not isinstance(placement_raw, dict):
                    raise ValueError(f"Each placement in {bed_name} / {phase_name} must be an object.")
                crop = str(placement_raw.get("crop", "")).strip()
                if not crop:
                    raise ValueError(f"Each placement in {bed_name} / {phase_name} must include crop.")
                code = str(placement_raw.get("code", "")).strip().upper()
                if code:
                    if code in used_codes and code_map.get(crop) != code:
                        raise ValueError(f"Crop code conflict for {crop}: {code}")
                    used_codes.add(code)
                else:
                    if crop in code_map:
                        code = code_map[crop]
                    else:
                        code = generate_code(crop, used_codes)
                code_map[crop] = code

                color = placement_raw.get("color")
                if color:
                    color_hex = normalize_hex_color(str(color))
                    color_map[crop] = color_hex
                else:
                    color_hex = assign_color(crop, color_map)

                x = int(placement_raw["x"])
                y = int(placement_raw["y"])
                w = int(placement_raw.get("w", 1))
                h = int(placement_raw.get("h", 1))
                if min(x, y, w, h) <= 0:
                    raise ValueError(f"Placement values must be positive in {bed_name} / {phase_name}.")
                if x + w - 1 > cols or y + h - 1 > rows:
                    raise ValueError(
                        f"Placement {crop} exceeds bounds in {bed_name} / {phase_name}."
                    )

                for row_offset in range(h):
                    for col_offset in range(w):
                        key = (x + col_offset, y + row_offset)
                        if key in occupied:
                            raise ValueError(
                                f"Overlapping placements in {bed_name} / {phase_name} at cell {key}."
                            )
                        occupied[key] = crop

                placement = {
                    "crop": crop,
                    "label": str(placement_raw.get("label", crop)).strip() or crop,
                    "code": code,
                    "x": x,
                    "y": y,
                    "w": w,
                    "h": h,
                    "color": color_hex,
                    "variety": str(placement_raw.get("variety", "")).strip(),
                    "notes": str(placement_raw.get("notes", "")).strip(),
                }
                placements.append(placement)

                if crop not in legend:
                    legend[crop] = {
                        "crop": crop,
                        "code": code,
                        "color": color_hex,
                        "cells": 0,
                        "placements": 0,
                        "locations": set(),
                    }
                legend[crop]["cells"] += w * h
                legend[crop]["placements"] += 1
                legend[crop]["locations"].add(f"{bed_name} / {phase_name}")

            phases.append(
                {
                    "name": phase_name,
                    "notes": str(phase_raw.get("notes", "")).strip(),
                    "placements": placements,
                }
            )

        beds.append(
            {
                "name": bed_name,
                "width_in": width_in,
                "length_in": length_in,
                "cols": cols,
                "rows": rows,
                "notes": str(bed_raw.get("notes", "")).strip(),
                "phases": phases,
            }
        )

    return {
        "title": title,
        "assumptions": assumptions,
        "beds": beds,
        "legend": legend,
    }


def make_grid_border() -> Border:
    grey = Side(style="thin", color="BFBFBF")
    return Border(left=grey, right=grey, top=grey, bottom=grey)


def make_block_border(is_top: bool, is_bottom: bool, is_left: bool, is_right: bool) -> Border:
    medium = Side(style="medium", color="7F7F7F")
    thin = Side(style="thin", color="BFBFBF")
    return Border(
        left=medium if is_left else thin,
        right=medium if is_right else thin,
        top=medium if is_top else thin,
        bottom=medium if is_bottom else thin,
    )


def style_sheet_grid(ws, rows: int, cols: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    grid_border = make_grid_border()

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row=GRID_START_ROW, column=GRID_START_COL)
    ws.page_setup.scale = 100
    ws.page_setup.orientation = "landscape" if cols >= rows else "portrait"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = f"$1:${HEADER_ROW}"
    ws.print_title_cols = "$A:$A"

    ws.column_dimensions["A"].width = 10
    for col_index in range(cols):
        column_letter = get_column_letter(GRID_START_COL + col_index)
        ws.column_dimensions[column_letter].width = GRID_COLUMN_WIDTH
        header_cell = ws.cell(row=HEADER_ROW, column=GRID_START_COL + col_index)
        header_cell.value = (col_index + 1) * CELL_SIZE_INCHES
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment

    for row_index in range(rows):
        grid_row = GRID_START_ROW + row_index
        ws.row_dimensions[grid_row].height = GRID_ROW_HEIGHT_POINTS
        row_header = ws.cell(row=grid_row, column=1)
        row_header.value = (row_index + 1) * CELL_SIZE_INCHES
        row_header.fill = header_fill
        row_header.font = header_font
        row_header.alignment = header_alignment
        for col_index in range(cols):
            grid_cell = ws.cell(row=grid_row, column=GRID_START_COL + col_index)
            grid_cell.border = grid_border
            grid_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def render_summary(ws, spec: dict) -> None:
    title_font = Font(size=16, bold=True)
    section_font = Font(size=12, bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    ws.title = "Summary"
    ws["A1"] = spec["title"]
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {datetime.now().isoformat(timespec='minutes')}"
    ws["A3"] = "Each plotted grid cell = 2 in x 2 in."

    current_row = 5
    if spec["assumptions"]:
        ws.cell(row=current_row, column=1, value="Assumptions").font = section_font
        current_row += 1
        for assumption in spec["assumptions"]:
            ws.cell(row=current_row, column=1, value=f"- {assumption}")
            current_row += 1
        current_row += 1

    ws.cell(row=current_row, column=1, value="Beds").font = section_font
    current_row += 1
    bed_headers = ["Bed", "Width (in)", "Length (in)", "Grid", "Phases"]
    for index, header in enumerate(bed_headers, start=1):
        cell = ws.cell(row=current_row, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    current_row += 1
    for bed in spec["beds"]:
        ws.cell(row=current_row, column=1, value=bed["name"])
        ws.cell(row=current_row, column=2, value=bed["width_in"])
        ws.cell(row=current_row, column=3, value=bed["length_in"])
        ws.cell(row=current_row, column=4, value=f"{bed['cols']} x {bed['rows']} cells")
        ws.cell(
            row=current_row,
            column=5,
            value=", ".join(phase["name"] for phase in bed["phases"]),
        )
        current_row += 1

    current_row += 1
    ws.cell(row=current_row, column=1, value="Legend").font = section_font
    current_row += 1
    legend_headers = ["Code", "Crop", "Color", "Placements", "Occupied cells", "Locations"]
    for index, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=current_row, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    current_row += 1

    for crop_name in sorted(spec["legend"]):
        item = spec["legend"][crop_name]
        ws.cell(row=current_row, column=1, value=item["code"])
        ws.cell(row=current_row, column=2, value=item["crop"])
        color_cell = ws.cell(row=current_row, column=3, value=f"#{item['color']}")
        color_cell.fill = PatternFill("solid", fgColor=item["color"])
        color_cell.font = Font(color=contrast_font_color(item["color"]), bold=True)
        ws.cell(row=current_row, column=4, value=item["placements"])
        ws.cell(row=current_row, column=5, value=item["cells"])
        ws.cell(row=current_row, column=6, value=", ".join(sorted(item["locations"])))
        current_row += 1

    widths = {
        "A": 16,
        "B": 28,
        "C": 14,
        "D": 14,
        "E": 16,
        "F": 48,
    }
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def render_bed_phase_sheet(ws, spec: dict, bed: dict, phase: dict) -> None:
    ws["A1"] = spec["title"]
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Bed: {bed['name']}"
    ws["A3"] = f"Phase: {phase['name']}"
    ws["A4"] = f"Dimensions: {int(bed['width_in'])} in x {int(bed['length_in'])} in ({bed['cols']} x {bed['rows']} cells)"
    if bed["notes"] or phase["notes"]:
        ws["D2"] = f"Bed notes: {bed['notes']}" if bed["notes"] else ""
        ws["D3"] = f"Phase notes: {phase['notes']}" if phase["notes"] else ""
        ws.column_dimensions["D"].width = 36
        ws.column_dimensions["E"].width = 36

    style_sheet_grid(ws, bed["rows"], bed["cols"])

    for placement in phase["placements"]:
        fill = PatternFill("solid", fgColor=placement["color"])
        font = Font(bold=True, size=10, color=contrast_font_color(placement["color"]))
        start_row = GRID_START_ROW + placement["y"] - 1
        start_col = GRID_START_COL + placement["x"] - 1
        center_row = start_row + ((placement["h"] - 1) // 2)
        center_col = start_col + ((placement["w"] - 1) // 2)
        comment_lines = [placement["crop"]]
        if placement["variety"]:
            comment_lines.append(f"Variety: {placement['variety']}")
        if placement["notes"]:
            comment_lines.append(placement["notes"])
        comment_text = "\n".join(comment_lines)

        for row_offset in range(placement["h"]):
            for col_offset in range(placement["w"]):
                is_top = row_offset == 0
                is_bottom = row_offset == placement["h"] - 1
                is_left = col_offset == 0
                is_right = col_offset == placement["w"] - 1
                cell = ws.cell(row=start_row + row_offset, column=start_col + col_offset)
                cell.fill = fill
                cell.border = make_block_border(is_top, is_bottom, is_left, is_right)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.value = placement["label"]
                cell.font = font

        center_cell = ws.cell(row=center_row, column=center_col)
        if comment_text:
            from openpyxl.comments import Comment

            center_cell.comment = Comment(comment_text, "Codex")

    max_col = GRID_START_COL + bed["cols"] - 1
    max_row = GRID_START_ROW + bed["rows"] - 1
    ws.print_area = f"$A$1:${get_column_letter(max_col)}${max_row}"


def build_workbook(spec: dict) -> tuple[Workbook, dict]:
    workbook = Workbook()
    render_summary(workbook.active, spec)
    existing_titles = set(workbook.sheetnames)
    workbook_summary = {"title": spec["title"], "beds": [], "sheets": ["Summary"]}

    for bed in spec["beds"]:
        bed_summary = {
            "name": bed["name"],
            "width_in": bed["width_in"],
            "length_in": bed["length_in"],
            "grid_cols": bed["cols"],
            "grid_rows": bed["rows"],
            "phases": [],
        }
        for phase in bed["phases"]:
            sheet_name = safe_sheet_title(f"{bed['name']} - {phase['name']}", existing_titles)
            ws = workbook.create_sheet(sheet_name)
            render_bed_phase_sheet(ws, spec, bed, phase)
            workbook_summary["sheets"].append(sheet_name)
            bed_summary["phases"].append(
                {
                    "name": phase["name"],
                    "sheet": sheet_name,
                    "placements": len(phase["placements"]),
                }
            )
        workbook_summary["beds"].append(bed_summary)

    return workbook, workbook_summary


def main() -> int:
    args = parse_args()
    spec_path = Path(args.spec)
    output_path = Path(args.output)
    spec = load_spec(spec_path)
    workbook, summary = build_workbook(spec)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    summary["output"] = str(output_path)
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
